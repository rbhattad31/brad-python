from openpyxl import *
import os
from openpyxl.chart import BarChart3D, Reference, series
# creates a new workbook
from openpyxl.utils.cell import col

wb = Workbook()
# Gets the first active worksheet
ws = wb.active
# creating new worksheets by using the create_sheet method

ws1 = wb.create_sheet("sheet1", 0)
ws2 = wb.create_sheet("sheet2", 1)

# Renaming the sheet
ws.title = "sample_1"

# save the workbook
if os.path.isfile("sample_1.xlsx"):
    pass
else:
    wb.save(filename="sample_1.xlsx")

# write headings

ws['A1'] = "Number1"
ws['B1'] = "Number2"
ws['C1'] = "Number3"

# write to file
rows = ((88, 46, 57), (89, 38, 12), (23, 59, 78), (56, 21, 98), (24, 18, 43), (34, 15, 67))
for row in rows:
    ws.append(row)
wb.save("sample_1.xlsx")

# merge cells
ws.merge_cells('B2:C5')
ws['B2'] = "Merged cells"
wb.save("sample_1.xlsx")

# read file
wb = load_workbook("sample_1.xlsx")
print("sheetnames: ", wb.sheetnames)
# max row
m_r = ws.max_row
# min column
m_c = ws.max_column
# Iterate sheet value
for value in ws.iter_rows(min_row=1, max_row=m_r, min_col=1, max_col=m_c, values_only=True):
    print(value)

# read data specific column
first_column = ws['A']

# Print the contents
for x in range(len(first_column)):
    print(first_column[x].value)





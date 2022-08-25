import openpyxl

# Read config details and parse to dictionary
def create_config(config_path, config_sheet):
    config = {}
    try:
        work_book = openpyxl.load_workbook(config_path)
        work_sheet = work_book[config_sheet]
        maximum_row = work_sheet.max_row
        maximum_col = work_sheet.max_column
        for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            cell_Name = config_details[0].value
            cell_value = config_details[1].value
            config[cell_Name] = cell_value
        return config
    
    except Exception as config_error:
        print("Purchase register Process failed to load config file. Hence stopping the BOT")
        return config_error

# Driver code
path = r"C:\Users\Hi\Python Projects\L & Co\Input\Config.xlsx"
create_config(path, "Concentration")

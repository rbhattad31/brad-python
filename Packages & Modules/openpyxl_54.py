# Append values
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

data = (
    (11, 48, 50),
    (81, 30, 82),
    (20, 51, 72),
    (21, 14, 60),
    (28, 41, 49),
    (74, 65, 53),
    ("johny", 'Depp', 45.63)
)

for i in data:
    sheet.append(i)
wb.save('values.xlsx')
#  Read Data from cell
import openpyxl

wb = openpyxl.load_workbook('values.xlsx')

sheet = wb.active

x1 = sheet['A1']
x2 = sheet['A2']
# using cell() function
x3 = sheet.cell(row=3, column=1)

print("The first cell value:", x1.value)
print("The second cell value:", x2.value)
print("The third cell value:", x3.value)


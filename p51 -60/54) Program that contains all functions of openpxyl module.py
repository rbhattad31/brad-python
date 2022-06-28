from openpyxl import Workbook
import time

#   Creating workbook

book = Workbook()

#   Creating worksheet

sheet = book.active

#   Input data to Excel cells of worksheet

sheet['A1'] = 56
sheet['A2'] = 43

now = time.strftime("%x")
sheet['A3'] = now

#   Row and Column notation method input

sheet.cell(row=2, column=2).value = 2   # B2 = 2


rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)  # append values to worksheet

#   read cell

a1 = sheet['A1']
a2 = sheet['A2']
print(a1.value)
print(a2.value)

#   freeze pane

sheet.freeze_panes = 'B2'

#   merge cells

sheet.merge_cells('A1:B2')

#   save workbook

book.save("sample.xlsx")

#   print the sheet names

print(book.sheetnames)

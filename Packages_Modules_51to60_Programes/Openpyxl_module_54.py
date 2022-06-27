# Append values
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active
data = (
    (1, 4, 5),
    (8, 3, 8),
    (2, 5, 7),
    (2, 1, 6),
    (2, 1, 4),
    (7, 5, 3),
    ("johny", 'Depp', 45.63)
)
for i in data:
    sheet.append(i)
wb.save('values.xlsx')

#  Read Data from cell

import openpyxl

wb = openpyxl.load_workbook('values.xlsx')
sheet = wb.active
x1 = sheet['A1']
x2 = sheet['A2']

# using cell() function
x3 = sheet.cell(row=4, column=3)

print("The first cell value:", x1.value)
print("The second cell value:", x2.value)
print("The third cell value:", x3.value)

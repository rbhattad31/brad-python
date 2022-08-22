import openpyxl
import Sourcecode.Comparatives.Purchase_type_wise_comparatives as ptcomp
import Sourcecode.Comparatives.Month_Wise_Comparatives as mwcomp
import Sourcecode.Comparatives.Plant_wise_comparatives as pwcomp
import Sourcecode.Comparatives.DomImp_wise_sheet as dicomp
import Sourcecode.Concentrations.PurcahseTypewise as ptconc
import Sourcecode.Concentrations.Monthwise as mwconc
import Sourcecode.Concentrations.Plantwise as pwconc
import Sourcecode.Concentrations.DomesicandImporttypewise as diconc


print("After importing modules")


# function "reading_sheets_from_config" reads the config file and creates a dictionary that contains Sheet name keys and
# Sheet name values that are next used to read sheet wise config data
def reading_sheets_from_config():
    try:
        config_sheets = {}
        work_book = openpyxl.load_workbook(r'Input\Config.xlsx')
        work_sheet = work_book["Sheet1"]
        maximum_row = work_sheet.max_row
        maximum_col = work_sheet.max_column

        for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            cell_Name = config_details[0].value
            cell_value = config_details[1].value
            config_sheets[cell_Name] = cell_value

        for key, value in config_sheets.items():
            print(key, ' : ', value)

        return config_sheets

    except Exception as config_error:
        print("Purchase register Concentration failed to load config file. Hence stopping the BOT")
        print(config_error)


# function "reading_sheet_config_data_to_dict" reads sheet wise config file and creates sheet specific config dictionary
def reading_sheet_config_data_to_dict(sheet_name):
    try:
        config = {}
        work_book = openpyxl.load_workbook(r'Input\Config.xlsx')
        work_sheet = work_book[sheet_name]
        maximum_row = work_sheet.max_row
        maximum_col = work_sheet.max_column

        for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            cell_Name = config_details[0].value
            cell_value = config_details[1].value
            config[cell_Name] = cell_value

        for key, value in config.items():
            print(key, ' : ', value)

        return config

    except Exception as config_error:
        print("Purchase register Concentration failed to load config file. Hence stopping the BOT")
        print(config_error)


config_sheets_dict = reading_sheets_from_config()

print("comparatives code execution is started")

config_comparatives = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Comparatives_sheetname"])
print("Executing Comparatives Purchase type code")
ptcomp.create_purchase_type_wise(config_comparatives)
print("Executing Comparatives Month wise code")
mwcomp .generate_month_wise(config_comparatives)
print("Executing Comparatives Plant wise code")
pwcomp.create_plant_wise_sheet(config_comparatives)
print("Executing Comparatives Domestic and Import wise code")
dicomp.generate_domestic_and_import_wise(config_comparatives)
print("comparatives code execution is complete")

# Concentrations code
print("Concentration code starts now")
# Read config details and parse to dictionary
config_concentration = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Concentration_sheetname"])

print("Executing concentration Purchase type code")
ptconc.purchasetype(config_concentration)
print("Executing concentration Month wise code")
mwconc.PurchaseMonth(config_concentration)
print("Executing concentration Plant wise code")
pwconc.PurchaseType(config_concentration)
print("Executing concentration Domestic and Import wise code")
diconc.PurchaseType(config_concentration)

print("Concentration code execution is Complete")

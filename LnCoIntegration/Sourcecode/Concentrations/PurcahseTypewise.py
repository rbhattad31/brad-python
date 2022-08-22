import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_lowercase
from openpyxl.styles import Alignment


class BusinessException(Exception):
    pass


def purchasetype(in_config):

    try:

        # Read Purchase Register Sheets
        Q4Sheet = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"], header=in_config["Q4 Header"])
        print("Dataframe Created")

        # Check Exception
        if Q4Sheet.shape[0] == 0:
            raise BusinessException("Sheet is empty")
        elif len(Q4Sheet["Valuation Class"]) == 0:
            raise BusinessException("Valuation Class Column is empty")
        elif len(Q4Sheet["Valuation Class Text"]) == 0:
            raise BusinessException("Valuation Class Text Column is empty")
        elif len(Q4Sheet["GR Amt.in loc.cur."]) == 0:
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # create Pivot Table Q4
        pivot_index = ["Valuation Class", "Valuation Class Text"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_Q4 = pd.pivot_table(Q4Sheet, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                  margins_name='Grand Total', sort=True)
        print("Q4 Pivot Table Created")

        # Remove Index
        pivot_Q4 = pivot_Q4.reset_index()

        # Assign Sheet
        pivot_sheet = pivot_Q4

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row of Q4 and Q3 columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[2]] == 0)], inplace=True)

        # Create Variance Column
        pivot_sheet['Variance'] = ""

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 2]

        # Variance Formula
        for index in pivot_sheet.index:
            quarter_value = pivot_sheet[col_name[2]][index]
            if total_value == 0:
                variance = 1
            else:
                variance = quarter_value / total_value

            pivot_sheet['Variance'][index] = variance

        print("Variance Column Created")

        # Change Column names of Q4 and Q3
        pivot_sheet = pivot_sheet.rename(columns={col_name[2]: in_config["Q4 Column"]})

        # Log Sheet
        pivot_sheet.to_excel(in_config["Type_Path"], sheet_name=in_config["TypeSheet"], index=False)
        print("Type Wise Comparatives Logged")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(in_config["Type_Path"])
        ws = wb[in_config["TypeSheet"]]

        # Format Q3
        for cell in ws['C']:
            cell.number_format = "#,###,##"

        # Format Variance
        for cell in ws['D']:
            cell.number_format = '0.0%'

        # Format Header
        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        for c in ascii_lowercase:
            ws[c + "1"].font = format_font

        # Format Footer
        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = format_font

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "1"].fill = format_fill
            if c == 'd':
                break

        # Footer Fill
        for c in ascii_lowercase:
            ws[c + str(m_row)].fill = format_fill
            if c == 'd':
                break

        # Merge Grand Total
        ws.merge_cells('A'+str(m_row)+':B'+str(m_row))
        cell = ws['A'+str(m_row)]
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set Width
        for c in ascii_lowercase:
            ws.column_dimensions[c].width = 20

        # Save File
        wb.save(in_config["Type_Path"])
        return ws

    except PermissionError as file_error:
        return file_error
    except FileNotFoundError as notfound_error:
        return notfound_error
    except BusinessException as business_error:
        return business_error
    except ValueError as value_error:
        return value_error
    except TypeError as type_error:
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        return error
    except KeyError as key_error:
        return key_error


config = {}

if __name__ == "__main__":
    purchasetype(config)


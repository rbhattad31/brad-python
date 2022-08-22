import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase


class BusinessException(Exception):
    pass


def create_plant_wise_sheet(in_config):
    try:
        read_excel_data = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"], skiprows=6)

        # Check Exception
        if read_excel_data.shape[0] == 0:
            raise BusinessException("Sheet is empty")
        elif len(read_excel_data["Plant"]) == 0:
            raise BusinessException("Plant Column is empty")
        elif len(read_excel_data["GR Amt.in loc.cur."]) == 0:
            raise BusinessException("GR Amt Column is empty")
        else:
            pass


        pivot_1 = pd.pivot_table(read_excel_data, index=["Plant"],
                                 values='GR Amt.in loc.cur.',
                                 aggfunc=numpy.sum, margins=True, margins_name='Grand Total')
        pivot_1 = pivot_1.reset_index()



        # read previous quarters final working file
        pivot_2 = pd.read_excel(in_config["Q3Path"], sheet_name=in_config["Q3 Sheet_Plant"],
                                                       usecols="A,C")



        merge_pd = pd.merge(pivot_1, pivot_2, how="outer", on=["Plant"])

        merge_pd = merge_pd.replace(numpy.nan, 0, regex=True)


        col_name = merge_pd.columns.values.tolist()

        # deleting columns present and past quarters both have values as zero
        merge_pd.drop(merge_pd.index[(merge_pd[col_name[1]] == 0) & (merge_pd[col_name[2]] == 0)],
                                         inplace=True)


        # creating a column in our output excel file
        merge_pd['Variance'] = ""

        pd.options.mode.chained_assignment = None

        # variance formula for index
        for index in merge_pd.index:
            Q4 = merge_pd[col_name[1]][index]
            Q3 = merge_pd[col_name[2]][index]

            if Q3 == 0:
                variance = 1
            else:
                variance = (Q4 - Q3) / Q3
            merge_pd['Variance'][index] = variance



        plant_wise_comparative_file = merge_pd.rename(
            columns={col_name[1]: in_config["Q4 Column"]})

        plant_wise_comparative_file = plant_wise_comparative_file.rename(
            columns={col_name[2]: in_config["Q3 Column"]})

        plant_wise_comparative_file.to_excel(in_config["Plant_Path"], sheet_name=in_config["PlantSheet"], index=False)


        wb = openpyxl.load_workbook(in_config["Plant_Path"])
        ws = wb[in_config["PlantSheet"]]

        for cell in ws['B']:
            cell.number_format = '#,###.##'
        for cell in ws['C']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '0.0%'

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style

        m_row = ws.max_row
        # Footer
        for i in ascii_uppercase:
            ws[i + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")

        for j in ascii_uppercase:
            ws[j + "1"].fill = fill_pattern
            if j == 'D':
                break

        for k in ascii_uppercase:
            ws.column_dimensions[k].width = 20

        ws.column_dimensions["D"].width = 12

        wb.save(in_config["Plant_Path"])

        return plant_wise_comparative_file

    # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except PermissionError as file_error:
        return file_error
    except OSError as error:
        return error
    except ImportError as I_error:
        return I_error
    except MemoryError as M_error:
        return M_error
    except RuntimeError as R_error:
        return R_error
    except TypeError as t_error:
        return t_error
    except BusinessException as business_error:
        return business_error
    except Exception as E_error:
        return E_error


config = {}


if __name__ == "__main__":
    print(create_plant_wise_sheet(config))

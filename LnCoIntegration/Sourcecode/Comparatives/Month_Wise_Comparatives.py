import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase


class BusinessException(Exception):
    pass


def generate_month_wise(in_config):
    try:
        read_excel_pd = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"], skiprows=6)

        # Check Exception
        if read_excel_pd.shape[0] == 0:
            raise BusinessException("Sheet is empty")
        elif len(read_excel_pd["Month"]) == 0:
            raise BusinessException("Month Column is empty")
        elif len(read_excel_pd["GR Amt.in loc.cur."]) == 0:
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        sort_order = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'September': 9,
                      'October': 10, 'November': 11, 'December': 12, 'Grand Total': 13}

        pivot_1 = pd.pivot_table(read_excel_pd, index=["Month"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum,
                                       margins=True, margins_name='Grand Total', sort=False)
        # reset index
        pivot_Q4 = pivot_1.reset_index()

        # Sort based on month
        month_wise_pd = pivot_Q4.sort_values('Month', key=lambda x: x.apply(lambda y: sort_order[y]))
        month_wise_pd.reset_index(inplace=True, drop=True)
        # Remove Empty Rows
        pivot_sheet = month_wise_pd.replace(numpy.nan, '', regex=True)


        # read from previous quarters final working file
        previous_quarter_final_file_pd = pd.read_excel(in_config["Q3Path"],
                                                       sheet_name=in_config["Q3 Sheet_Month"],
                                                       usecols="C, D")

        # concatenation instead of merge as there are no common Columns to merge.
        concat_pd = pd.concat([pivot_sheet, previous_quarter_final_file_pd], axis=1)

        concat_pd = concat_pd.replace(numpy.nan, 0, regex=True)

        columns_names = concat_pd.columns.values.tolist()
        # returns as ['Month', 'GR Amt.in loc.cur.', 'Month.1', 'Previous Quarter as Q3 FY 21-22']

        concat_pd.drop(concat_pd.index[(concat_pd[columns_names[1]] == 0) & (concat_pd[columns_names[3]] == 0)],
                                        inplace=True)

        # create a new column - Success
        concat_pd['Variance'] = ''

        pd.options.mode.chained_assignment = None

        # variance formula implementation using index
        for index in concat_pd.index:
            Q4 = concat_pd[columns_names[1]][index]
            Q3 = concat_pd[columns_names[3]][index]

            if Q3 == 0:
                variance = 1
            else:
                variance = (Q4 - Q3) / Q3

            concat_pd['Variance'][index] = variance

        month_wise_comparatives_pd = concat_pd.rename(
            columns={columns_names[1]: in_config["Q4 Column"]})

        month_wise_comparatives_pd = month_wise_comparatives_pd.rename(
            columns={columns_names[2]: in_config["R_Month"]})

        month_wise_comparatives_pd.to_excel(in_config["Month_Path"], sheet_name=in_config["MonthSheet"],
                                            index=False)
        wb = openpyxl.load_workbook(in_config["Month_Path"])
        ws = wb[in_config["MonthSheet"]]

        for cell in ws['B']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '#,###.##'
        for cell in ws['E']:
            cell.number_format = '0%'

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style

        # Footer
        m_row = ws.max_row

        for i in ascii_uppercase:
            ws[i + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")

        for j in ascii_uppercase:
            ws[j + "1"].fill = fill_pattern
            if j == 'E':
                break

        for k in ascii_uppercase:
            ws.column_dimensions[k].width = 20

        wb.save(in_config["Month_Path"])

        return month_wise_comparatives_pd

    # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except PermissionError as file_error:
        return file_error
    except OSError as error:
        return error
    except ImportError as I_error:
        return I_error
    except MemoryError as M_error:
        return M_error
    except RuntimeError as R_error:
        return R_error
    except TypeError as t_error:
        return t_error
    except BusinessException as business_error:
        return business_error
    except Exception as E_error:
        return E_error

config = {}


if __name__ == "__main__":
    print(generate_month_wise(config))

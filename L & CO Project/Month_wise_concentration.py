import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase


def create_month_wise(Input_Excel_data_file):
    try:
        #  selecting only required columns
        read_excel_pd = pd.read_excel(Input_Excel_data_file,
                                 sheet_name="Purchase register Q4", skiprows=6)

        sort_order = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'September': 9,
                      'October': 10, 'November': 11, 'December': 12, 'Grand Total': 13}


        pivot_1 = pd.pivot_table(read_excel_pd, index=["Month"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                  margins_name='Grand Total')


        # Remove Index
        pivot_Q4 = pivot_1.reset_index()

        # Sort based on month
        month_wise_pd = pivot_Q4.sort_values('Month', key=lambda x: x.apply(lambda y: sort_order[y]))
        month_wise_pd.reset_index(inplace=True, drop=True)



        # Remove Empty Rows
        pivot_sheet = month_wise_pd.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row of Q4 and Q3 columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[1]] == 0)])

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 1]

        # Variance Formula
        variance_list = []
        for index in pivot_sheet.index:
            quarter_value = (pivot_sheet[col_name[1]][index])

            if total_value == 0:
                variance = 1
            else:
                variance = quarter_value / total_value
            variance_list.append(variance)
        pivot_sheet['Variance']= variance_list





        pivot_sheet.to_excel("month_concentration.xlsx", sheet_name="Month Wise Concentration",
                                              index=False)


        wb = openpyxl.load_workbook("month_concentration.xlsx")
        ws = wb["Month Wise Concentration"]

        for cell in ws['B']:
            cell.number_format = '#,###'
        for cell in ws['C']:
            cell.number_format = '0%'

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style

        # Footer
        m_row = ws.max_row
        for i in ascii_uppercase:
            ws[i + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")

        for j in ascii_uppercase:
            ws[j + "1"].fill = fill_pattern
            if j == 'C':
                break

        for k in ascii_uppercase:
            ws.column_dimensions[k].width = 20

        wb.save("month_concentration.xlsx")


        return pivot_sheet
    # Excepting Errors here
        # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except TypeError as t_error:
        return t_error
create_month_wise("purchase registers.xlsx")

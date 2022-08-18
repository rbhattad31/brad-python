import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase

class BusinessException(Exception):
    pass

def generate_domestic_and_import_wise(in_config):
    try:
        read_excel_pd = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"], skiprows=6)

        # Check Exception
        if read_excel_pd.shape[0] == 0:
            raise BusinessException("Sheet is empty")
        elif len(read_excel_pd["GR Amt.in loc.cur."]) == 0:
            raise BusinessException("GR Amt Column is empty")
        else:
            pass


        # create a new column 'Purchase Type' with blank value
        read_excel_pd['Purchase Type'] = ''

        # Setting Type of purchase column values using currency key column on condition
        read_excel_pd.loc[read_excel_pd['Currency Key'] == "INR", 'Purchase Type'] = "Domestic"
        read_excel_pd.loc[read_excel_pd['Currency Key'] != "INR", 'Purchase Type'] = "Import"

        #  selecting only required columns
        read_excel_pd = read_excel_pd[['Purchase Type', 'GR Amt.in loc.cur.']]

        # create pivot table - sorting not required
        domestic_and_import_wise_pd = pd.pivot_table(read_excel_pd, index=["Purchase Type"],
                                                     values="GR Amt.in loc.cur.",
                                                     aggfunc=numpy.sum, margins=True, margins_name="Grand Total")


        # reset month index after pivot table creation for concatenation
        domestic_and_import_wise_pd = domestic_and_import_wise_pd.reset_index()

        # read previous quarters final working file
        previous_quarter_final_file_pd = pd.read_excel(in_config["Q3Path"],
                                                       sheet_name=in_config["Q3 Sheet_Dom&Imp"],
                                                       usecols="A,C")


        # merging present and previous quarter purchase type wise data
        merge_pd = pd.merge(domestic_and_import_wise_pd, previous_quarter_final_file_pd,
                            how="outer", on=["Purchase Type"])

        columns_list = merge_pd.columns.values.tolist()
        # create a new column - Success
        merge_pd['Variance'] = 0

        # To Remove SettingWithCopyWarning error
        pd.options.mode.chained_assignment = None  # modifying only one df, so suppressing this warning as it is not affecting

        # variance formula for index
        for index in merge_pd.index:
            Q4 = merge_pd[columns_list[1]][index]
            Q3 = merge_pd[columns_list[2]][index]

            if Q3 == 0:
                variance = 1
            else:
                variance = (Q4 - Q3) / Q3
            merge_pd['Variance'][index] = variance

        domestic_and_import_wise_comparatives_pd = merge_pd.rename(
            columns={columns_list[1]: in_config["Q4 Column"]})

        domestic_and_import_wise_comparatives_pd = domestic_and_import_wise_comparatives_pd.rename(
            columns={columns_list[2]: in_config["Q3 Column"]})

        domestic_and_import_wise_comparatives_pd.to_excel(in_config["Dom&Imp_Path"],
                                                          sheet_name=in_config["Dom&Imp Sheet"], index=False)

        wb = openpyxl.load_workbook(in_config["Dom&Imp_Path"])
        ws = wb[in_config["Dom&Imp Sheet"]]

        for cell in ws['B']:
            cell.number_format = '#,###.##'
        for cell in ws['C']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '0.0%'

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "1"].font = font_style

        m_row = ws.max_row
        for c in ascii_uppercase:
            ws[c + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")
        for c in ascii_uppercase:
            ws[c + "1"].fill = fill_pattern
            if c == 'D':
                break

        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 20

        wb.save(in_config["Dom&Imp_Path"])

        return domestic_and_import_wise_comparatives_pd

    # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except PermissionError as file_error:
        return file_error
    except OSError as error:
        return error
    except ImportError as I_error:
        return I_error
    except MemoryError as M_error:
        return M_error
    except RuntimeError as R_error:
        return R_error
    except TypeError as t_error:
        return t_error
    except BusinessException as business_error:
        return business_error
    except Exception as E_error:
        return E_error
config = {}
try:
    work_book = openpyxl.load_workbook(r"C:\Users\Hello\OneDrive\Desktop\config.xlsx")
    work_sheet = work_book["Comparatives"]
    maximum_row = work_sheet.max_row
    maximum_col = work_sheet.max_column
    for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
        cell_Name = config_details[0].value
        cell_value = config_details[1].value
        config[cell_Name] = cell_value

except Exception as config_error:
    print("Purchase register Process failed to load config file. Hence stopping the BOT")
    print(config_error)

print(generate_domestic_and_import_wise(config))

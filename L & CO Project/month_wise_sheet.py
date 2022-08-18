import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase


def generate_month_wise(Input_file, previous_quarter_final_file):
    try:
        read_excel_1 = pd.read_excel(Input_file, sheet_name="Purchase register Q4", skiprows=6)

        sort_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        month_wise_pd = pd.pivot_table(read_excel_1, index=["Month"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum,
                                       margins=True, margins_name='Grand Total', sort=False)

        month_wise_pd.to_excel("excel_1.xlsx", sheet_name="Sheet1")

        read_excel_2 = pd.read_excel("excel_1.xlsx", sheet_name="Sheet1")

        read_excel_2.index = pd.CategoricalIndex(read_excel_2['Month'], categories=sort_order, ordered=True)
        month_wise_pd = read_excel_2.sort_index().reset_index(drop=True)

        month_wise_pd.to_excel("excel_2.xlsx", sheet_name="Sheet1", index=False)

        month_file = pd.read_excel("excel_2.xlsx", sheet_name="Sheet1")

        # read from previous quarters final working file
        previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file,
                                                       sheet_name="Month Wise Comparitives",
                                                       usecols=(2, 3))

        # concatenation instead of merge as there are no common Columns to merge.
        month_wise_comparatives_pd = pd.concat([month_file, previous_quarter_final_file_pd], axis=1)

        month_wise_comparatives_pd = month_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)

        columns_names = month_wise_comparatives_pd.columns.values.tolist()
        # returns as ['Month', 'GR Amt.in loc.cur.', 'Month.1', 'Previous Quarter as Q3 FY 21-22']

        month_wise_comparatives_pd.drop(month_wise_comparatives_pd.index[
                                            (month_wise_comparatives_pd[columns_names[1]] == 0) & (
                                                    month_wise_comparatives_pd[
                                                        columns_names[3]] == 0)],
                                        inplace=True)

        # create a new column - Success
        month_wise_comparatives_pd['Variance'] = ''

        pd.options.mode.chained_assignment = None  # modifying only one df, so suppressing this warning as it is not affecting

        # variance formula implementation using index
        for index in month_wise_comparatives_pd.index:
            present_quarter = month_wise_comparatives_pd[columns_names[1]][index]
            previous_quarter = month_wise_comparatives_pd[columns_names[3]][index]

            if previous_quarter == 0:
                variance = 1
            else:
                variance = (present_quarter - previous_quarter) / previous_quarter

            month_wise_comparatives_pd['Variance'][index] = variance

        month_wise_comparatives_pd = month_wise_comparatives_pd.rename(
            columns={columns_names[1]: "Q4 FY 21-22"})

        month_wise_comparatives_pd = month_wise_comparatives_pd.rename(
            columns={columns_names[2]: "Month"})

        month_wise_comparatives_pd.to_excel("month_wise_comparatives_1.xlsx", sheet_name="Month Wise Comparatives",
                                            index=False)
        wb = openpyxl.load_workbook("month_wise_comparatives_1.xlsx")
        ws = wb["Month Wise Comparatives"]

        for cell in ws['B']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '#,###.##'
        for cell in ws['E']:
            cell.number_format = '0%'

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style

        # Footer
        m_row = ws.max_row

        for i in ascii_uppercase:
            ws[i + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")

        for j in ascii_uppercase:
            ws[j + "1"].fill = fill_pattern
            if j == 'E':
                break

        for k in ascii_uppercase:
            ws.column_dimensions[k].width = 20

        wb.save("month_wise_comparatives_1.xlsx")

        return month_wise_comparatives_pd
    # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except TypeError as t_error:
        return t_error


generate_month_wise("purchase registers.xlsx", "sample.xlsx")

import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase




def create_plant_wise_sheet(Past_Quarter_File):
    try:

        read_excel_pd = pd.read_excel(Past_Quarter_File, sheet_name="Purchase register Q4", skiprows=6)

        pivot_1 = pd.pivot_table(read_excel_pd, index=["Plant"],
                                 values='GR Amt.in loc.cur.',
                                 aggfunc=numpy.sum, margins=True, margins_name='Grand Total')

        plant_wise_comparative_file= pivot_1.reset_index()


        col_name = plant_wise_comparative_file.columns.values.tolist()

        # deleting columns present and past quarters both have values as zero
        plant_wise_comparative_file.drop(plant_wise_comparative_file.index[
                                             (plant_wise_comparative_file[col_name[1]] == 0)],
                                         inplace=True)
        # creating a column in our output excel file
        plant_wise_comparative_file['Variance'] = ""

        pd.options.mode.chained_assignment = None

        Grand_total_value = plant_wise_comparative_file[col_name[1]].iloc[-1]

        for index in plant_wise_comparative_file.index:
            row_value = plant_wise_comparative_file[col_name[1]][index]
            variance_1 = row_value / Grand_total_value
            plant_wise_comparative_file['Variance'][index] = variance_1


        plant_wise_comparative_file = plant_wise_comparative_file.rename(
            columns={col_name[1]: "Q4 FY 21-22"})



        plant_wise_comparative_file.to_excel("Plant_wise_concentration.xlsx", sheet_name="Plant Wise Concentration", index=False)

        wb = openpyxl.load_workbook("Plant_wise_concentration.xlsx")
        ws = wb["Plant Wise Concentration"]

        for cell in ws['B']:
            cell.number_format = '#,###.##'

        for cell in ws['C']:
            cell.number_format = '0%'

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style


        m_row = ws.max_row
        # Footer
        for i in ascii_uppercase:
            ws[i + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")

        for j in ascii_uppercase:
            ws[j + "1"].fill = fill_pattern
            if j == 'C':
                break

        for k in ascii_uppercase:
            ws.column_dimensions[k].width = 20


        wb.save("Plant_wise_concentration.xlsx")

        return plant_wise_comparative_file
    # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except TypeError as t_error:
        return t_error


create_plant_wise_sheet("purchase registers.xlsx")

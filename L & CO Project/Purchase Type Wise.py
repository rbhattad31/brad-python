# Importing Libraries

import pandas as pd
import numpy
import timeit
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_lowercase



# Execution time starts here
start_time = timeit.default_timer()


# Defining a Function


def Create_Purchase_Type_Wise_File(Input_Excel_file, Past_Registered_Quarter_File):
    try:

        r1_excel = pd.read_excel(Input_Excel_file, sheet_name="Purchase register Q4", skiprows=6)
        r2_excel = pd.read_excel(Past_Registered_Quarter_File, sheet_name="Purchase register Q3", skiprows=5)

        pivot_1 = pd.pivot_table(r1_excel, index=["Valuation Class", "Valuation Class Text"],
                                 values='GR Amt.in loc.cur.',
                                 aggfunc=numpy.sum, margins=True, margins_name='Grand Total', sort=True)
        pivot_1 = pivot_1.reset_index()

        pivot_2 = pd.pivot_table(r2_excel, index=["Valuation Class", "Valuation Class Text"],
                                 values='GR Amt.in loc.cur.',
                                 aggfunc=numpy.sum, margins=True, margins_name='Grand Total', sort=True)
        pivot_2 = pivot_2.reset_index()


        purchase_type_wise_comparatives_File = pd.merge(pivot_1, pivot_2, how="outer",
                                                        on=["Valuation Class", "Valuation Class Text"])

        purchase_type_wise_comparatives_File = purchase_type_wise_comparatives_File.replace(numpy.nan, 0, regex=True)

        # return as List
        col_name = purchase_type_wise_comparatives_File.columns.values.tolist()

        # deleting columns present and past quarters both have values as zero
        purchase_type_wise_comparatives_File.drop(purchase_type_wise_comparatives_File.index[
                                                      (purchase_type_wise_comparatives_File[col_name[2]] == 0) & (
                                                              purchase_type_wise_comparatives_File[
                                                                  col_name[3]] == 0)],
                                                  inplace=True)
        # creating a column in our output excel file
        purchase_type_wise_comparatives_File['Variance'] = ""

        pd.options.mode.chained_assignment = None


        # variance formula for index
        for index in purchase_type_wise_comparatives_File.index:
            present_quarter = purchase_type_wise_comparatives_File[col_name[2]][index]
            past_quarter = purchase_type_wise_comparatives_File[col_name[3]][index]

            if past_quarter == 0:
                variance = 1
            else:
                variance = (present_quarter - past_quarter) / past_quarter
            purchase_type_wise_comparatives_File['Variance'][index] = variance



        # copy present quarter Amount column Grand total, set it as zero, sort the data frame and reassign the value.
        grand_total = purchase_type_wise_comparatives_File[col_name[2]].values[-1]
        purchase_type_wise_comparatives_File[col_name[2]].values[-1] = 0


        purchase_type_wise_comparatives_File[col_name[2]].values[-1] = grand_total

        purchase_type_wise_comparatives_File = purchase_type_wise_comparatives_File.rename(
            columns={col_name[2]: "Q4 FY 21-22"})
        purchase_type_wise_comparatives_File = purchase_type_wise_comparatives_File.rename(
            columns={col_name[3]: "Q3 FY 21-22"})



        purchase_type_wise_comparatives_File.to_excel("Purchase_type_wise_comparatives.xlsx", sheet_name="Purchase Type Wise Comparatives",
                                                      index=False)


        wb = openpyxl.load_workbook("Purchase_type_wise_comparatives.xlsx")
        ws= wb["Purchase Type Wise Comparatives"]

        for cell in ws['C']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '#,###.##'
        for cell in ws['E']:
            cell.number_format = '0.0%'



        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for c in ascii_lowercase:
            ws[c + "1"].font = font_style

        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")
        for c in ascii_lowercase:
            ws[c + "1"].fill = fill_pattern
            if c == 'e':
                break



        for c in ascii_lowercase:
            ws.column_dimensions[c].width = 25
        ws.column_dimensions["E"].width = 15

        wb.save("Purchase_type_wise_comparatives.xlsx")

        return purchase_type_wise_comparatives_File


   # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except TypeError as t_error:
        return t_error

result = Create_Purchase_Type_Wise_File(Input_Excel_file="purchase registers.xlsx",
                                        Past_Registered_Quarter_File="purchase registers.xlsx")



# Execution time ends here
end_time = timeit.default_timer()
# Calculating Execution time in seconds
execution_time = (end_time - start_time)
# Print Execution time
print("Program is Executed in seconds: ", execution_time)

import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from string import ascii_uppercase



def create_plant_wise_sheet(Present_Quarter_File):
    try:

        read_excel_pd = pd.read_excel(Present_Quarter_File, sheet_name="Purchase register Q4", skiprows=6)

        pivot_1 = pd.pivot_table(read_excel_pd, index=["Valuation Class", "Valuation Class Text"],
                                 values='GR Amt.in loc.cur.',
                                 aggfunc=numpy.sum, margins=True, margins_name='Grand Total')
        purchase_wise_concentration_pd = pivot_1.reset_index()




        col_name = purchase_wise_concentration_pd.columns.values.tolist()
        # deleting columns present and past quarters both have values as zero
        purchase_wise_concentration_pd.drop(
            purchase_wise_concentration_pd.index[(purchase_wise_concentration_pd[col_name[2]] == 0)],
            inplace=True)

        # creating a column in our output excel file
        purchase_wise_concentration_pd['Variance'] = ""
        pd.options.mode.chained_assignment = None

        Grand_total_value = purchase_wise_concentration_pd[col_name[2]].iloc[-1]

        for index in purchase_wise_concentration_pd.index:
            row_value = purchase_wise_concentration_pd[col_name[2]][index]
            variance_1 = row_value / Grand_total_value
            purchase_wise_concentration_pd['Variance'][index] = variance_1

        plant_wise_concentration_pd = purchase_wise_concentration_pd.rename(
            columns={col_name[2]: "Q4 FY 21-22"})

        plant_wise_concentration_pd.to_excel("Purchase_concentration.xlsx", sheet_name="Purchase Type Wise",
                                             index=False)
        wb = openpyxl.load_workbook("Purchase_concentration.xlsx")
        ws = wb["Purchase Type Wise"]

        for cell in ws['C']:
            cell.number_format = '#,###'
        for cell in ws['D']:
            cell.number_format = '0%'

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style

        # Footer
        m_row = ws.max_row

        for i in ascii_uppercase:
            ws[i + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")

        for j in ascii_uppercase:
            ws[j + "1"].fill = fill_pattern
            if j == 'D':
                break

        for k in ascii_uppercase:
            ws.column_dimensions[k].width = 25

        ws.column_dimensions["D"].width = 14

        ws.merge_cells('A11:B11')
        cell = ws['A11']
        cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save("Purchase_concentration.xlsx")

        return plant_wise_concentration_pd

    # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except TypeError as t_error:
        return t_error



create_plant_wise_sheet(Present_Quarter_File="purchase registers.xlsx")

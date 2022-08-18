import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase


def generate_domestic_and_import_wise(excel_file_pd):
    try:
        read_excel_pd = pd.read_excel(excel_file_pd, sheet_name="Purchase register Q4", skiprows=6)

        read_excel_pd['Purchase Type'] = ''

        # Setting Type of purchase column values using currency key column on condition
        read_excel_pd.loc[read_excel_pd['Currency Key'] == "INR", 'Purchase Type'] = "Domestic"
        read_excel_pd.loc[read_excel_pd['Currency Key'] != "INR", 'Purchase Type'] = "Import"

        # create pivot table - sorting not required
        domestic_and_import_wise_pd = pd.pivot_table(read_excel_pd, index=["Purchase Type"],
                                                     values="GR Amt.in loc.cur.",
                                                     aggfunc=numpy.sum, margins=True, margins_name="Grand Total")

        # reset month index after pivot table creation for concatenation
        domestic_and_import_concentration_pd = domestic_and_import_wise_pd.reset_index()

        columns_list = domestic_and_import_concentration_pd.columns.values.tolist()
        # deleting columns present and past quarters both have values as zero
        domestic_and_import_concentration_pd.drop(domestic_and_import_concentration_pd.index[
                                                      (domestic_and_import_concentration_pd[columns_list[1]] == 0)],
                                                  inplace=True)

        # create a new column - Success
        domestic_and_import_concentration_pd['Variance'] = 0

        # To Remove SettingWithCopyWarning error
        pd.options.mode.chained_assignment = None

        Grand_total_value = domestic_and_import_concentration_pd[columns_list[1]].iloc[-1]

        for index in domestic_and_import_concentration_pd.index:
            row_value = domestic_and_import_concentration_pd[columns_list[1]][index]

            variance_1 = row_value / Grand_total_value
            domestic_and_import_concentration_pd['Variance'][index] = variance_1

        domestic_and_import_concentration_pd = domestic_and_import_concentration_pd.rename(
            columns={columns_list[1]: "Q4 FY 21-22"})

        domestic_and_import_concentration_pd.to_excel("Dom&Imp_concentration.xlsx",
                                                      sheet_name="Dom&Imp Wise Concentration", index=False)

        wb = openpyxl.load_workbook("Dom&Imp_concentration.xlsx")
        ws = wb["Dom&Imp Wise Concentration"]

        for cell in ws['B']:
            cell.number_format = '#,###.##'

        for cell in ws['C']:
            cell.number_format = '0.0%'

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "1"].font = font_style

        # Footer
        m_row = ws.max_row
        for c in ascii_uppercase:
            ws[c + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="8080FF")
        for c in ascii_uppercase:
            ws[c + "1"].fill = fill_pattern
            if c == 'C':
                break

        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 20

        wb.save("Dom&Imp_concentration.xlsx")
        return domestic_and_import_concentration_pd

    # Excepting Errors here
    except FileNotFoundError as error:
        return error
    except ValueError as V_error:
        return V_error
    except KeyError as k_error:
        return k_error
    except NameError as n_error:
        return n_error
    except TypeError as t_error:
        return t_error


generate_domestic_and_import_wise("purchase registers.xlsx")

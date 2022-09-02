from string import ascii_lowercase
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import PatternFill
from win32com import client
import pywintypes
import os


class BusinessException(Exception):
    pass


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def VendorNumbers_Duplication(in_config):
    try:
        # Read Purchase Register Sheets
        vendor_data = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Sheet"])
        print("Dataframe Created")

        # Fetch To Address
        to_address = in_config["To_Address"]
        cc_address = in_config["CC_Address"]

        # Check data in input sheet
        if vendor_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        vendor_col = vendor_data.columns.values.tolist()
        for col in ["Vendor", "Name 1", "Tax Number"]:
            if col not in vendor_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter rows
        vendor_name = vendor_data[vendor_data['Name 1'].notna()]
        vendor_no = vendor_data[vendor_data['Vendor'].notna()]
        vendor_tax = vendor_data[vendor_data['Tax Number'].notna()]

        # Check Exception
        if len(vendor_no) == 0:
            subject = in_config["EmptyVendorNo_Subject"]
            body = in_config["EmptyVendorNo_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Vendor Number Column is empty")
        elif len(vendor_name) == 0:
            subject = in_config["EmptyVendorName_Subject"]
            body = in_config["EmptyVendorName_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Vendor Name Column is empty")
        elif len(vendor_tax) == 0:
            subject = in_config["EmptyTax_Subject"]
            body = in_config["EmptyTax_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Tax Number Column is empty")
        else:
            pass

        # Mark Empty rows
        vendor_data = vendor_data.replace(numpy.nan, "Empty", regex=True)

        # create Pivot Table
        pivot_index = ["Vendor", "Name 1", "Tax Number"]
        pivot_data = pd.pivot_table(vendor_data, index=pivot_index, sort=True)
        print("Pivot Table Created")

        # Remove Index
        pivot_data = pivot_data.reset_index()

        # Assign Sheet
        pivot_sheet = pivot_data[["Vendor", "Name 1", "Tax Number"]]

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)
        pivot_sheet = pivot_sheet.replace("Empty", '', regex=True)

        # Create Duplicate Column
        pivot_sheet['Duplicate'] = ""
        pivot_sheet['Lower case'] = pivot_sheet["Name 1"].str.lower()

        # Map Duplicate Rows
        pivot_sheet['Duplicate'] = pivot_sheet.duplicated(subset=["Lower case"], keep=False) \
            .map({True: 'Yes', False: 'No'})

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Assign Sheet
        pivot_sheet = pivot_sheet[["Vendor", "Name 1", "Tax Number", 'Duplicate']]

        # Delete row where vendor number columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[0]] == 0)], inplace=True)

        # Sort Table
        pivot_sheet = pivot_sheet.sort_values(by='Vendor', ascending=True)

        # Log Sheet
        pivot_sheet.to_excel(in_config["DuplicationResult_Path"], sheet_name=in_config["DuplicationSheet"], index=False)

        # Check outfile creation
        if os.path.exists(in_config["DuplicationResult_Path"]):
            print("Duplication of vendor numbers logged successfully")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(in_config["DuplicationResult_Path"])
        ws = wb[in_config["DuplicationSheet"]]

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "1"].fill = format_fill
            if c == 'c':
                break

        # Highlight row
        format_fill = PatternFill(patternType='solid', fgColor='FFFF00')
        for cell in ws['d']:
            if cell.value == 'Yes':
                ws['B' + str(cell.row)].fill = format_fill

        # Set Width
        for c in ascii_lowercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'c':
                break

        # Delete Duplicate Column
        ws.delet_col(idx=4)

        # Save File
        wb.save(in_config["DuplicationResult_Path"])
        return ws

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        return notfound_error
    except BusinessException as business_error:
        print("Duplication Process-", end="")
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        return value_error
    except TypeError as type_error:
        cc
        print("Duplication Process-", end="")
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Duplication Process-", end="")
        print("Please check the given keyword is correct")
        return key_error


# Read config details and parse to dictionary
config = {}
try:
    work_book = openpyxl.load_workbook(r'Input/Duplication Config.xlsx')
    work_sheet = work_book["Duplication"]
    maximum_row = work_sheet.max_row
    maximum_col = work_sheet.max_column
    for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
        cell_Name = config_details[0].value
        cell_value = config_details[1].value
        config[cell_Name] = cell_value

except Exception as config_error:
    print("Duplication of vendor process failed to load config file. Hence stopping the BOT")
    print(config_error)

print(VendorNumbers_Duplication(config))

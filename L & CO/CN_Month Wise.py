import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_lowercase
from win32com import client
import pywintypes
import os


class BusinessException(Exception):
    pass


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def PurchaseMonth(in_config):
    try:
        # Read Purchase Register Sheets
        Q4Sheet = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"], header=in_config["Q4 Header"])
        print("Dataframe Created")

        # Fetch To Address
        to_address = in_config["To_Address"]
        cc_address = in_config["CC_Address"]

        # Check Exception
        if Q4Sheet.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        Q4Sheet_col = Q4Sheet.columns.values.tolist()
        for col in ['GR Posting Date', "GR Amt.in loc.cur."]:
            if col not in Q4Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter rows
        date_1 = Q4Sheet[Q4Sheet['GR Posting Date'].notna()]
        gr_amt = Q4Sheet[Q4Sheet['GR Amt.in loc.cur.'].notna()]

        if len(date_1) == 0:
            subject = in_config["Date_Subject"]
            body = in_config["Date_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Month Column is empty")
        elif len(gr_amt) == 0:
            subject = in_config["GRAmt_Subject"]
            body = in_config["GRAmt_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # Create Month Column
        Q4Sheet['GR Posting Date'] = pd.to_datetime(Q4Sheet['GR Posting Date'])
        Q4Sheet['GR Posting Date'] = Q4Sheet['GR Posting Date'].dt.month_name().str[:3]
        Q4Sheet['Month'] = Q4Sheet['GR Posting Date']

        # Sort based on month
        month_dict = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9,
                      'Oct': 10, 'Nov': 11, 'Dec': 12, 'Grand Total': 13}

        # Create Pivot Table Q3
        pivot_index = ["Month"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_Q4 = pd.pivot_table(Q4Sheet, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                  margins_name='Grand Total')

        # Get Pivot Column Names
        col_name = pivot_Q4.columns.values.tolist()

        # Rename Column
        pivot_Q4 = pivot_Q4.rename(columns={col_name[0]: in_config["Q4 Column"]})

        # Remove Index
        pivot_Q4 = pivot_Q4.reset_index()

        # Sort based on month
        pivot_Q4 = pivot_Q4.sort_values('Month', key=lambda x: x.apply(lambda x: month_dict[x]))
        pivot_Q4.reset_index(inplace=True, drop=True)
        print("Q4 Pivot Table Created")

        # Assign Sheets
        pivot_sheet = pivot_Q4

        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row of Q4 and Q3 columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[1]] == 0)])

        pd.options.mode.chained_assignment = None

        # Get maximum value
        total_value = pivot_sheet.iloc[-1:]
        total_value = total_value.iloc[0, 1]

        # Variance Formula
        variance_list = []
        for index in pivot_sheet.index:
            quarter_value = (pivot_sheet[col_name[1]][index])

            if total_value == 0:
                variance = 1
            else:
                variance = quarter_value / total_value

            variance_list.append(variance)

        # Create Variance Column
        pivot_sheet['Variance'] = variance_list
        print("Variance Column Created")

        # Log Sheet
        pivot_sheet.to_excel(in_config["Month_Path"], sheet_name=in_config["MonthSheet"], index=False)

        # Check outfile creation
        if os.path.exists(in_config["Month_Path"]):
            print("Month Wise Concentration Logged")
        else:
            subject = in_config["OutputNotFound_Subject"]
            body = in_config["OutputNotFound_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Output file not generated")

        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(in_config["Month_Path"])
        ws = wb[in_config["MonthSheet"]]

        # Format Q3
        for cell in ws['B']:
            cell.number_format = "#,###,##"

        # Format Variance
        for cell in ws['C']:
            cell.number_format = '0.0%'

        # Format Header
        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        for c in ascii_lowercase:
            ws[c + "1"].font = format_font

        # Format Footer
        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = format_font

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "1"].fill = format_fill
            if c == 'c':
                break

        # Footer Fill
        for c in ascii_lowercase:
            ws[c + str(m_row)].fill = format_fill
            if c == 'c':
                break

        # Set Width
        for c in ascii_lowercase:
            ws.column_dimensions[c].width = 15

        # Save File
        wb.save(in_config["Month_Path"])
        return ws

    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", end="")
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", end="")
        return notfound_error
    except BusinessException as business_error:
        print("Concentration Month Wise Process-", end="")
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", end="")
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", end="")
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", end="")
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Concentration Month Wise Process-", end="")
        return key_error


# Read config details and parse to dictionary
config = {}
try:
    work_book = openpyxl.load_workbook(r'../Input/Config.xlsx')
    work_sheet = work_book["CN_MonthWise"]
    maximum_row = work_sheet.max_row
    maximum_col = work_sheet.max_column
    for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
        cell_Name = config_details[0].value
        cell_value = config_details[1].value
        config[cell_Name] = cell_value

except Exception as config_error:
    print("Purchase register Process failed to load config file. Hence stopping the BOT")
    print(config_error)

print(PurchaseMonth(config))

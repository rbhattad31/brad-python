import sys
import openpyxl
import pywintypes
from win32com import client


import Sourcecode.Comparatives.Purchase_type_wise_comparatives as ptcomp
import Sourcecode.Comparatives.Month_Wise_comparatives as mwcomp
import Sourcecode.Comparatives.Plant_wise_comparatives as pwcomp
import Sourcecode.Comparatives.DomImp_wise_sheet_comparatives as dicomp
import Sourcecode.Concentrations.PurcahseTypewiseConcentration as ptconc
import Sourcecode.Concentrations.MonthWiseConcentration as mwconc
import Sourcecode.Concentrations.PlantWiseConcentration as pwconc
import Sourcecode.Concentrations.DomAndImportConcentration as diconc
import Sourcecode.DuplicationofVendorNumbers as duplication
import Sourcecode.averagedaypurchase as averagedaypurchase
import Sourcecode.SameMaterialPurchasesfromDVDP as smpdvdp
import Sourcecode.Unit_Price_Comparsion as upc


print("Process is started")


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def reading_sheets_names_from_config_main_sheet():
    try:
        config_sheets = {}
        work_book = openpyxl.load_workbook(r'C:\Users\Hello\PycharmProjects\L & CO Final Demo to Client\Input\Config.xlsx')
        work_sheet = work_book["Main"]
        maximum_row = work_sheet.max_row
        maximum_col = work_sheet.max_column

        for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            cell_Name = config_details[0].value
            cell_value = config_details[1].value
            config_sheets[cell_Name] = cell_value

        return config_sheets

    except Exception as config_error:
        print("failed to load config file. Hence stopping the BOT")
        print(config_error)
        to = "pavankumar.pamidi@bradsol.com"
        cc = "pavankumar.pamidi@bradsol.com"
        subject = "Failed to read Config File and Fetching sheet names"
        body = '''
Hello,

Config file is failed to load, hence Bot stopped processing. 

Thanks & Regards,
L & Co  
'''
        send_mail(to=to, cc=cc, subject=subject, body=body)
        sys.exit(1)


# function "reading_sheet_config_data_to_dict" reads sheet wise config file and creates sheet specific config dictionary
def reading_sheet_config_data_to_dict(sheet_name):
    try:
        config = {}
        work_book = openpyxl.load_workbook(r"C:\Users\Hello\PycharmProjects\L & CO Final Demo to Client\Input\Config.xlsx")
        work_sheet = work_book[sheet_name]
        maximum_row = work_sheet.max_row
        maximum_col = work_sheet.max_column

        for config_details in work_sheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            cell_Name = config_details[0].value
            cell_value = config_details[1].value
            config[cell_Name] = cell_value

        return config

    except Exception as config_error:
        print("failed to load config file for sheet:", sheet_name)
        print(config_error)
        to = "pavankumar.pamidi@bradsol.com"
        cc = "pavankumar.pamidi@bradsol.com"
        subject = "Config reading is failed for sheet: " + sheet_name
        body = '''
Hello,

Config file is failed to load. Continuing with next process.

Thanks & Regards,
L & Co  

'''
        send_mail(to=to, cc=cc, subject=subject, body=body)
        raise Exception


print("Reading main sheet config file is started")
config_sheets_dict = reading_sheets_names_from_config_main_sheet()
print("Reading main sheet config file is completed")

print("*******************************************")
# send Bot starting mail
start_to = config_sheets_dict['To_Address']
start_cc = config_sheets_dict['CC_Address']
start_subject = config_sheets_dict['Start_Mail_Subject']
start_body = config_sheets_dict['Start_Mail_Body']
send_mail(to=start_to, cc=start_cc, body=start_body, subject=start_subject)
print("Process start mail notification is sent")

print("*******************************************")

print("Executing Comparatives Purchase type code")
try:
    config_purchase_comparatives = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Comparatives_Purchase_sheetname"])
    ptcomp.create_purchase_type_wise(config_purchase_comparatives)
except Exception as e:
    print("Exception caught for Process: Purchase type comparatives, Continuing with next sheet process")

print("*******************************************")

print("Executing Comparatives Month wise code")
try:
    config_Month_comparatives = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Comparatives_Month_sheetname"])
    mwcomp.purchasemonth(config_Month_comparatives)
except Exception as e:
    print("Exception caught for Process: Month wise comparatives, Continuing with next sheet process")

print("*******************************************")

print("Executing Comparatives Plant wise code")
try:
    config_plant_comparatives = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Comparatives_Plant_sheetname"])
    pwcomp.create_plant_wise_sheet(config_plant_comparatives)
except Exception as e:
    print("Exception caught for Process: Plant wise comparatives, Continuing with next sheet process")

print("*******************************************")

print("Executing Comparatives Domestic and Import wise code")
try:
    config_domandimp_comparatives = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Comparatives_Dom&Imp_sheetname"])
    dicomp.generate_domestic_and_import_wise(config_domandimp_comparatives)
except Exception as e:
    print("Exception caught for Process: Domestic and import wise comparators, continuing with next sheet process")

print("*******************************************")

print("Executing concentration Purchase type code")
try:
    config_concentration_ptw = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Concentrations_Purchase_sheetname"])
    ptconc.PurchaseType(config_concentration_ptw)
except Exception as e:
    print("Exception caught for Process: Purchase type concentration, continuing with next sheet process")

print("*******************************************")

print("Executing concentration Month wise code")
try:
    config_concentration_mw = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Concentrations_Month_sheetname"])
    mwconc.month_wise(config_concentration_mw)
except Exception as e:
    print("Exception caught for Process: Month wise Concentration, Continuing with next sheet process")

print("*******************************************")

print("Executing concentration Plant wise code")
try:
    config_concentration_pw = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Concentrations_Plant_sheetname"])
    pwconc.PurchaseType(config_concentration_pw)
except Exception as e:
    print("Exception caught for Process: Plant wise concentration, continuing with next sheet process")

print("*******************************************")

print("Executing concentration Domestic and Import wise code")
try:
    config_concentration_DI = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Concentrations_Dom&Imp_sheetname"])
    diconc.purchase_type(config_concentration_DI)
except Exception as e:
    print("Exception caught for Process: domestic and import wise concentration")

print("*******************************************")

print("Executing Duplication of Vendor number code")
try:
    config_duplication = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Duplication_of_Vendor_sheetname"])
    duplication.vendor_numbers_duplication(config_duplication)
except Exception as e:
    print("Exception caught for Process: duplication of Vendor number")

print("*******************************************")

print("Executing Average Day Purchase code")
try:
    config_average_day_purchase = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Average_Day_Purchase_sheetname"])
    averagedaypurchase.average_day_purchase(config_average_day_purchase)
except Exception as e:
    print("Exception caught for Process: Average day purchase")
print("*******************************************")

print("Executing 'Same Material Purchases from Different Vendors & Different prices' code")
try:
    config_smpdvdp = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Same_Material_Purchases_DVDP_sheetname"])
    smpdvdp.same_mat_dvp(config_smpdvdp)
except Exception as e:
    print("Exception caught for Process: Same Material different Vendors")
print("*******************************************")

print("Executing Unit Price Comparison code")
try:
    config_unit_price = reading_sheet_config_data_to_dict(sheet_name=config_sheets_dict["Unit_Price_Comparison_sheetname"])
    upc.create_unit_price(config_unit_price)
except Exception as e:
    print("Exception caught for Process: Unit Price Comparison code")
print("*******************************************")

# Bot success mail notification
end_to = config_sheets_dict['To_Address']
end_cc = config_sheets_dict['CC_Address']
end_subject = config_sheets_dict['Success_Mail_Subject']
end_body = config_sheets_dict['Success_Mail_Body']
send_mail(to=end_to, cc=end_cc, body=end_body, subject=end_subject)
print("Process complete mail notification is sent")
print("Bot successfully finished Processing of the sheets")

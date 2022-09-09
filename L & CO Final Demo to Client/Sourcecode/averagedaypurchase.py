import os
from string import ascii_lowercase
import xlrd_compdoc_commented as xlrd
import openpyxl
import pywintypes
import pandas as pd
from win32com import client
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font


config = {}     # Empty Config dictionary to be used to assign key vale pairs


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as error_message:
        print("Sendmail error - Please check outlook connection")
        return error_message
    except Exception as error:
        return error

    finally:
        print('Mail process over')


#  Function produce average day purchase output sheet


def average_day_purchase(in_config):
    to_ = in_config["To_Address"]
    cc_ = in_config["CC_Address"]

    try:

        # Reading the data from excel
        dataframe = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config['Q4 Sheet'], skiprows=6)

        # Empty data check
        if dataframe.empty:
            print('Empty DataFrame')
            raise ValueError

        # Column existence check
        if pd.Series(['GR Amt.in loc.cur.', 'GR Posting Date']).isin(dataframe.columns).all():
            print("Columns exist")
        else:
            print("One or more columns doesnt exist in the dataframe")
            raise KeyError

        # Empty column data check
        if dataframe['GR Amt.in loc.cur.'].isnull().all():
            print('Column GR Amt.in loc.cur. is empty')
            raise ValueError

        if dataframe['GR Posting Date'].isnull().all():
            print('Column GR Posting Date is empty')
            raise ValueError

        # Column data format check
        if dataframe['GR Amt.in loc.cur.'].dtype == 'float64':
            print("Type is float")
        else:
            print("Column format not correct")
            raise TypeError

        if dataframe['GR Posting Date'].dtype == 'datetime64[ns]':
            print("Type is dd-mm-yyyy")
        else:
            print("Column format not correct")
            raise TypeError

        # Generating pivot table
        pivot_q4 = pd.pivot_table(dataframe, values='GR Amt.in loc.cur.', index='GR Posting Date',
                                  observed=True, sort=True, aggfunc='sum', margins=True, margins_name="Grand Total")

        # Sort pivot table
        pivot_q4 = pivot_q4.reset_index()



        # Store the last grand total value to variable
        total_gr_amt = (pivot_q4.iloc[-1]['GR Amt.in loc.cur.'])

        # Remove last row from pivot table
        pivot_q4 = pivot_q4.iloc[:-1]

        # Arithmetic operations
        pivot_q4['Average'] = total_gr_amt / (len(pivot_q4))
        pivot_q4['Variance'] = pivot_q4['Average'] - pivot_q4['GR Amt.in loc.cur.']
        pivot_q4['Percentage'] = pivot_q4['Variance'] / pivot_q4['Average']
        pivot_q4.columns = ['GR Posting Date', 'Amount as per purchase register', 'Average purchases', 'Variance',
                            'Percentage']

        # Sort pivot table based of percentage column in ascending order
        pivot_q4.sort_values(by='Percentage', inplace=True)

        # Save pivot table to Excel file
        pivot_q4.to_excel(in_config['output path'], sheet_name=in_config['new sheet'], startrow=4, index=False)

        #   Formatting and styling the Excel data

        # Load excel file
        workbook = openpyxl.load_workbook(in_config['output path'])

        # Load sheet
        worksheet = workbook[in_config['new sheet']]

        # Assign max row value to variable
        m_row = worksheet.max_row

        # Set column widths
        for c in ascii_lowercase:
            column_length = max(len(str(cell.value)) for cell in worksheet[c])
            worksheet.column_dimensions[c].width = column_length * 1.25
            if c == 'e':
                break

        # Implement subtotal formula for max row values
        worksheet['B4'] = '=SUBTOTAL(9,B6:B'+str(m_row)+')'
        worksheet['C4'] = '=SUBTOTAL(9,C6:C'+str(m_row)+')'

        # Set font style variable configuration
        font_style = Font(name='Cambria', size=11, color='002060', bold=True)

        # Implement the configuration to appropriate rows
        worksheet['B4'].font = font_style
        worksheet['C4'].font = font_style

        # Number format implementation
        for cell in worksheet['A']:
            cell.number_format = 'dd-mm-yyyy'

        for cell in worksheet['B']:
            cell.number_format = '#,###'

        for cell in worksheet['C']:
            cell.number_format = '#,###'

        for cell in worksheet['D']:
            cell.number_format = '#,###'

        for cell in worksheet['E']:
            cell.number_format = '000%'

        # Cell color variable assignment
        fill_cell = PatternFill(patternType='solid', fgColor='ffff00')
        fill_color = PatternFill(fgColor='d9e1f2', fill_type="solid")

        # Cell color implementation for appropriate rows
        for row in worksheet["A5:E5"]:
            for cell in row:
                cell.fill = fill_color
                cell.font = font_style

        for row in worksheet["A6:E9"]:
            for cell in row:
                cell.fill = fill_cell

        # Cell border implementation
        thin = Side(border_style="thin", color='b1c5e7')

        for row in worksheet.iter_rows(min_row=5, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row in worksheet.iter_rows(min_row=6, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.font = Font(name='Cambria', size=11, color='002060')

        for row in worksheet.iter_rows(min_row=4, min_col=1, max_row=4, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')

        workbook.save(in_config['output path'])

        # Empty cell check
        wb = xlrd.open_workbook(in_config['output path'])
        wb_sheet = wb.sheet_by_index(0)

        for row in range(5, wb_sheet.nrows):
            for column in range(0, wb_sheet.ncols):
                if wb_sheet.cell_value(row, column) == "":
                    print('row', row+1, 'col', column+1, 'is empty')
                    raise RuntimeError



    #  Exceptions handling

    except FileNotFoundError:
        print("Check with the file paths")
        print("Sending notification through outlook mail...")

        subject_ = 'Required file not found'
        body_ = in_config['FileNotFoundError']
        send_mail(to_, cc_, subject_, body_)

    except PermissionError:
        print("Close the excel file before execution")
        os.system('TASKKILL /F /IM excel.exe')
        print("Re executing...")
        average_day_purchase(config)

    except KeyError:
        print("Check with the input column names provided from the source file and program")
        print("Sending notification through outlook mail...")
        subject_ = 'Required column not found'
        body_ = in_config['KeyError']
        send_mail(to_, cc_, subject_, body_)

    except TypeError:
        print('Type error occurred')
        print("Sending notification through outlook mail...")
        subject_ = 'Incorrect column format'
        body_ = in_config['TypeError']
        send_mail(to_, cc_, subject_, body_)

    except RuntimeError:
        print('Run time error occurred')
        print("Sending notification through outlook mail...")
        subject_ = 'Empty cell found'
        body_ = in_config['RuntimeError']
        send_mail(to_, cc_, subject_, body_)

    except FileExistsError:
        print("Directory not found")
        print("Sending notification through outlook mail...")
        subject_ = 'Directory not found'
        body_ = in_config['FileExistsError']
        send_mail(to_, cc_, subject_, body_)

    except SystemError:
        print("System Exception occurred")

    except ArithmeticError:
        print("Look for overflow, zero division and floating point errors causing formula and values")

    except IndexError:
        print("Look for incorrect index values")

    except SyntaxError:
        print("Look for incorrect syntax input")

    finally:
        print("Process is over")

if __name__ == "__main__":
    print(average_day_purchase(config))


import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill
from string import ascii_uppercase

from win32com import client
import pywintypes

class BusinessException(Exception):
    pass
# Send Outlook Mails
def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error

def create_plant_wise_sheet(in_config):
    try:
        read_excel_data = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"], skiprows=6)
        read_excel_data_2 = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q3 Sheet_Plant"], skiprows=5)


        # Check Exception
        if read_excel_data.shape[0] == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")

        Q4Sheet_col = read_excel_data.columns.values.tolist()
        for col in ["Plant", "GR Amt.in loc.cur."]:
            if col not in Q4Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
                raise BusinessException(col + " Column is missing")
        # Filter Rows
        Plant_pd = read_excel_data[read_excel_data['Plant'].notna()]
        Gr_Amt_pd = read_excel_data[read_excel_data['GR Amt.in loc.cur.'].notna()]

        if len(Plant_pd) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Pant_subject"],
                      body=in_config["Plant_Body"])
            raise BusinessException("Plant Column is empty")

        elif len(Gr_Amt_pd) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass


        pivot_1 = pd.pivot_table(read_excel_data, index=["Plant"],
                                 values='GR Amt.in loc.cur.',
                                 aggfunc=numpy.sum, margins=True, margins_name='Grand Total')
        pivot_1 = pivot_1.reset_index()




        # read previous quarters final working file
        pivot_2 = pd.pivot_table(read_excel_data_2, index=["Plant"],
                                 values='GR Amt.in loc.cur.',
                                 aggfunc=numpy.sum, margins=True, margins_name='Grand Total')



        merge_pd = pd.merge(pivot_1, pivot_2, how="outer", on=["Plant"])

        merge_pd = merge_pd.replace(numpy.nan, 0, regex=True)


        col_name = merge_pd.columns.values.tolist()

        # deleting columns present and past quarters both have values as zero
        merge_pd.drop(merge_pd.index[(merge_pd[col_name[1]] == 0) & (merge_pd[col_name[2]] == 0)],
                                         inplace=True)


        # creating a column in our output excel file
        merge_pd['Variance'] = ""

        pd.options.mode.chained_assignment = None

        # variance formula for index
        for index in merge_pd.index:
            Q4 = merge_pd[col_name[1]][index]
            Q3 = merge_pd[col_name[2]][index]

            if Q3 == 0:
                variance = 1
            else:
                variance = (Q4 - Q3) / Q3
            merge_pd['Variance'][index] = variance



        plant_wise_comparative_file = merge_pd.rename(
            columns={col_name[1]: in_config["Q4 Column"]})

        plant_wise_comparative_file = plant_wise_comparative_file.rename(
            columns={col_name[2]: in_config["Q3 Column"]})

        plant_wise_comparative_file.to_excel(in_config["Plant_Path"], sheet_name=in_config["PlantSheet"], index=False)


        wb = openpyxl.load_workbook(in_config["Plant_Path"])
        ws = wb[in_config["PlantSheet"]]

        for cell in ws['B']:
            cell.number_format = '#,###.##'
        for cell in ws['C']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '0.0%'

        # Header
        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for i in ascii_uppercase:
            ws[i + "1"].font = font_style

        m_row = ws.max_row
        # Footer
        for i in ascii_uppercase:
            ws[i + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")

        for j in ascii_uppercase:
            ws[j + "1"].fill = fill_pattern
            if j == 'D':
                break

        for k in ascii_uppercase:
            ws.column_dimensions[k].width = 20

        ws.column_dimensions["D"].width = 12

        wb.save(in_config["Plant_Path"])

        return plant_wise_comparative_file

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Plant Type Wise Comparatives Process-", end="")
        return notfound_error
    except ValueError as V_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(V_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", end="")
        return V_error
    except BusinessException as business_error:
        print("Plant Type Wise Comparatives Process-", end="")
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Plant Type Wise Comparatives Process-", end="")
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Plant Type Wise Comparatives Process-", end="")
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Plant Type Wise Comparatives Process-", end="")
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Please close the file")
        return file_error


config = {}
if __name__ == "__main__":
    print(create_plant_wise_sheet(config))


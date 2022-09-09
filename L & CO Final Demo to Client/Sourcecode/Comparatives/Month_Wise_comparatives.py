import pandas as pd
import numpy
import openpyxl
from win32com import client
import pywintypes
from openpyxl.styles import Font, PatternFill
from string import ascii_lowercase


class BusinessException(Exception):
    pass
def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def purchasemonth(in_config):
    try:
        # Read Purchase Register Sheets
        Q4Sheet = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"], header=6)
        Q3Sheet = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q3 Sheet_Month"], header=5)



        # Check Exception
        if Q4Sheet.shape[0] == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")

        Q4Sheet_col = Q4Sheet.columns.values.tolist()
        for col in ["Month", "GR Amt.in loc.cur."]:
            if col not in Q4Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        Month_pd = Q4Sheet[Q4Sheet['Month'].notna()]
        Gr_Amt_pd = Q4Sheet[Q4Sheet['GR Amt.in loc.cur.'].notna()]

        if len(Month_pd) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Month_subject"],
                      body=in_config["Month_Body"])
            raise BusinessException("Month Column is empty")

        elif len(Gr_Amt_pd) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # create Pivot Table Q4
        pivot_index = ["Month"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_Q4 = pd.pivot_table(Q4Sheet, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                  margins_name='Grand Total')

        # Get Pivot Column Names
        col_name = pivot_Q4.columns.values.tolist()

        # Rename Column
        pivot_Q4 = pivot_Q4.rename(columns={col_name[0]: in_config["Q4 Column"]})
        pivot_Q4 = pivot_Q4.reset_index()

        # Sort based on month
        month_dict = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'September': 9,
                      'October': 10, 'November': 11, 'December': 12, 'Grand Total': 13}

        pivot_Q4 = pivot_Q4.sort_values('Month', key=lambda x: x.apply(lambda y: month_dict[y]))
        pivot_Q4.reset_index(inplace=True, drop=True)


        # Create Pivot Table Q3
        pivot_index = ["Month"]
        pivot_values = ["GR Amt.in loc.cur."]
        pivot_Q3 = pd.pivot_table(Q3Sheet, index=pivot_index, values=pivot_values, aggfunc=numpy.sum, margins=True,
                                  margins_name='Grand Total')

        # Get Pivot Column Names
        col_name = pivot_Q3.columns.values.tolist()

        # Rename Column
        pivot_Q3 = pivot_Q3.rename(columns={col_name[0]: in_config["Q3 Column"]})

        # Remove Index
        pivot_Q3 = pivot_Q3.reset_index()

        # Sort based on month
        pivot_Q3 = pivot_Q3.sort_values('Month', key=lambda x: x.apply(lambda a: month_dict[a]))
        pivot_Q3.reset_index(inplace=True, drop=True)


        # Merge Pivot Sheets
        pivot_sheet = pd.concat([pivot_Q4, pivot_Q3], axis=1, sort=False)


        # Remove Empty Rows
        pivot_sheet = pivot_sheet.replace(numpy.nan, '', regex=True)

        # Get Pivot Column Names
        col_name = pivot_sheet.columns.values.tolist()

        # Delete row of Q4 and Q3 columns values as zero
        pivot_sheet.drop(pivot_sheet.index[(pivot_sheet[col_name[1]] == 0) & (pivot_sheet[col_name[3]] == 0)])

        pd.options.mode.chained_assignment = None

        # Variance Formula
        variance_list = []
        for index in pivot_sheet.index:
            quarter_4 = (pivot_sheet[col_name[1]][index])
            quarter_3 = (pivot_sheet[col_name[3]][index])

            if quarter_3 == 0:
                variance = 1
            else:
                variance = (quarter_4 - quarter_3) / quarter_3

            variance_list.append(variance)

        # Create Variance Column
        pivot_sheet['Variance'] = variance_list


        # Log Sheet
        pivot_sheet.to_excel(in_config["Month_Path"], sheet_name=in_config["MonthSheet"], index=False)


        # Load Sheet in openpyxl
        wb = openpyxl.load_workbook(in_config["Month_Path"])
        ws = wb[in_config["MonthSheet"]]

        # Format Q4 & Q3
        for col in ['B', 'D']:
            for cell in ws[col]:
                cell.number_format = "#,###,##.##"

        # Format Variance
        for cell in ws['E']:
            cell.number_format = '0%'

        # Format Header
        format_font = Font(name="Calibri", size=11, color="000000", bold=True)
        for c in ascii_lowercase:
            ws[c + "1"].font = format_font

        # Format Footer
        m_row = ws.max_row
        for c in ascii_lowercase:
            ws[c + str(m_row)].font = format_font

        # Header Fill
        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "1"].fill = format_fill
            if c == 'e':
                break

        # Set Width
        for c in ascii_lowercase:
            ws.column_dimensions[c].width = 20

        # Save File
        wb.save(in_config["Month_Path"])
        return ws

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Month Type Wise Comparatives Process-", end="")
        return notfound_error
    except ValueError as V_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(V_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", end="")
        return V_error
    except BusinessException as business_error:
        print("Month Type Wise Comparatives Process-", end="")
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Month Type Wise Comparatives Process-", end="")
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Month Type Wise Comparatives Process-", end="")
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Month Type Wise Comparatives Process-", end="")
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Please close the file")
        return file_error


# Read config details and parse to dictionary
config = {}

if __name__ == "__main__":
    print(purchasemonth(config))


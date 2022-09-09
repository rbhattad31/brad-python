from string import ascii_lowercase

from win32com import client

import numpy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import numbers
import warnings
warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)

class BusinessException(Exception):
    pass


# Send Outlook Mails
def send_mail(to, cc, subject, body, pywintypes=None):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def create_unit_price(in_config):
    try:
        Excel_data = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"],
                                   skiprows=in_config["Skiprow_Q4"])

        # Fetch To Address
        to_address = in_config["To_Address"]
        cc_address = in_config["CC_Address"]

        # Check Exception
        if Excel_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        Q4Sheet_col = Excel_data.columns.values.tolist()
        for col in ["GR Amt.in loc.cur.", "GR Qty", "Material No.", "Valuation Class Text", "Vendor Name"]:
            if col not in Q4Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        pivot_Q4 = pd.pivot_table(Excel_data, index=["Material No.", "Valuation Class Text", "Vendor Name"],
                                  values=["GR Amt.in loc.cur.", "GR Qty"], aggfunc=numpy.sum, margins=True,
                                  margins_name="Grand Total", sort=True)
        # Drop last column of a dataframe
        pivot_Q4 = pivot_Q4[:-1]

        pivot_Q4 = pivot_Q4.reset_index()
        pivot_Q4 = pivot_Q4.replace(numpy.nan, 0, regex=True)
        columns = pivot_Q4.columns.values.tolist()
        numpy.seterr(divide='ignore')
        pivot_Q4['Unit Price'] = ""
        pd.options.mode.chained_assignment = None

        for index in pivot_Q4.index:
            GR_amt = pivot_Q4[columns[3]][index]
            GR_qty = pivot_Q4[columns[4]][index]
            Unit_price = GR_amt / GR_qty
            pivot_Q4['Unit Price'][index] = Unit_price

        columns = pivot_Q4.columns.values.tolist()
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[3]: "GR Amt.in loc.cur.1"})
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[4]: "GR Qty1"})
        pivot_Q4 = pivot_Q4.rename(
            columns={columns[5]: "Unit Price1"})

        pivot_Q4['Concat'] = ""
        pivot_Q4["Concat"] = pivot_Q4["Material No."].astype(str) + pivot_Q4["Valuation Class Text"] + pivot_Q4[
            "Vendor Name"]
        pivot_Q4 = pivot_Q4[["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1",
                             "GR Qty1", "Unit Price1"]]

        # Q3 Pivot
        Excel_data = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q3 Sheet"],
                                   skiprows=in_config["Skiprow_Q3"])

        # Check Exception
        if Excel_data.shape[0] == 0:
            subject = in_config["EmptyInput_Subject1"]
            body = in_config["EmptyInput_Body"]
            send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
            raise BusinessException("Sheet is empty")

        # Check Column Present
        Q3Sheet_col = Excel_data.columns.values.tolist()
        for col in ["GR Amt.in loc.cur.", "GR Qty", "Material No.", "Valuation Class Text", "Vendor Name"]:
            if col not in Q3Sheet_col:
                subject = in_config["ColumnMiss_Subject1"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)
                send_mail(to=to_address, cc=cc_address, subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        pivot_Q3_1 = pd.pivot_table(Excel_data, values=["GR Amt.in loc.cur.", "GR Qty"],
                                    index=["Material No.", "Valuation Class Text", "Vendor Name"],
                                    aggfunc={"GR Amt.in loc.cur.": numpy.sum, 'GR Qty': numpy.sum}, margins=True)

        pivot_Q3 = pivot_Q3_1
        pivot_Q3 = pivot_Q3.reset_index()
        pivot_Q3 = pivot_Q3.replace(numpy.nan, 0, regex=True)

        # Drop last column of a dataframe
        pivot_Q3 = pivot_Q3[:-1]
        columns = pivot_Q3.columns.values.tolist()

        pivot_Q3['Unit Price'] = ""
        pd.options.mode.chained_assignment = None

        for index in pivot_Q3.index:
            GR_amt = pivot_Q3[columns[3]][index]
            GR_qty = pivot_Q3[columns[4]][index]
            Unit_price = GR_amt / GR_qty
            pivot_Q3['Unit Price'][index] = Unit_price

        columns = pivot_Q3.columns.values.tolist()
        pivot_Q3 = pivot_Q3.rename(
            columns={columns[3]: "GR Amt.in loc.cur.2"})
        pivot_Q3 = pivot_Q3.rename(
            columns={columns[4]: "GR Qty2"})
        pivot_Q3 = pivot_Q3.rename(
            columns={columns[5]: "Unit Price2"})

        pivot_Q3['Concat'] = ""
        pivot_Q3["Concat"] = pivot_Q3["Material No."].astype(str) + pivot_Q3["Valuation Class Text"] + pivot_Q3[
            "Vendor Name"]
        pivot_Q3 = pivot_Q3[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.2", "GR Qty2",
             "Unit Price2"]]

        Unit_Price_Q4 = pd.merge(pivot_Q4, pivot_Q3, how="left",
                                 on=["Material No.", "Valuation Class Text", "Vendor Name", "Concat"], copy=False)
        Unit_Price_Q4 = Unit_Price_Q4[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1", "GR Qty1",
             "Unit Price1"]]
        Unit_Price_Q3 = pd.merge(pivot_Q3, pivot_Q4, how="left",
                                 on=["Material No.", "Valuation Class Text", "Vendor Name", "Concat"], copy=False)
        Unit_Price_Q3 = Unit_Price_Q3[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.2", "GR Qty2",
             "Unit Price2"]]
        columns = Unit_Price_Q3.columns.values.tolist()


        Unit_Price = pd.merge(Unit_Price_Q4, Unit_Price_Q3, how="outer",
                              on=["Material No.", "Valuation Class Text", "Vendor Name", "Concat"], copy=False)
        # Unit_Price = pd.concat([Unit_Price_Q4, Unit_Price_Q3], ignore_index=True, sort=True)
        Unit_Price = Unit_Price[
            ["Material No.", "Valuation Class Text", "Vendor Name", "Concat", "GR Amt.in loc.cur.1", "GR Qty1",
             "Unit Price1", "GR Amt.in loc.cur.2", "GR Qty2", "Unit Price2"]]
        Unit_Price = Unit_Price.reset_index()

        Unit_Price = Unit_Price.replace(numpy.nan, 0, regex=True)
        Unit_Price["Remarks"] = " "


        columns = Unit_Price.columns.values.tolist()
        columns.remove("Remarks")
        columns.insert(4, "Remarks")
        # Re-order columns
        Unit_Price = Unit_Price[columns]

        # Unit_Price_Comparison.sort_values
        columns = Unit_Price.columns.values.tolist()
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            if Unit_Price[columns[6]][index] == 0:
                Unit_Price['Remarks'][index] = "Not purchased in the current quarter"
            elif Unit_Price[columns[9]][index] == 0:
                Unit_Price['Remarks'][index] = "Purchased in the current quarter"
            else:
                pass

        Unit_Price['Increase/decrease in Amount'] = ""
        pd.options.mode.chained_assignment = None
        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[6]][index]
            GR_qty = Unit_Price[columns[9]][index]
            Unit_price = GR_amt - GR_qty
            Unit_Price['Increase/decrease in Amount'][index] = Unit_price

        Unit_Price['Increase/decrease in Quantity'] = ""
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[7]][index]
            GR_qty = Unit_Price[columns[10]][index]
            Unit_price = GR_amt - GR_qty
            Unit_Price['Increase/decrease in Quantity'][index] = Unit_price

        Unit_Price['Increase/decrease in Unit Price'] = ""
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[8]][index]
            GR_qty = Unit_Price[columns[11]][index]
            Unit_price = GR_amt - GR_qty
            Unit_Price['Increase/decrease in Unit Price'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['Increase/decrease in unit price (%)'] = ""
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[14]][index]
            GR_qty = Unit_Price[columns[11]][index]
            Unit_price = (GR_amt / GR_qty)
            Unit_Price['Increase/decrease in unit price (%)'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['In amount due to Qty'] = ""
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[13]][index]
            GR_qty = Unit_Price[columns[11]][index]
            Unit_price = GR_amt * GR_qty
            Unit_Price['In amount due to Qty'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['In amount due to Qty (%)'] = ""
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[16]][index]
            GR_qty = Unit_Price[columns[12]][index]
            Unit_price = GR_amt / GR_qty
            Unit_Price['In amount due to Qty (%)'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['In amount due to price'] = ""
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[7]][index]
            GR_qty = Unit_Price[columns[14]][index]
            Unit_price = GR_amt * GR_qty
            Unit_Price['In amount due to price'][index] = Unit_price

        columns = Unit_Price.columns.values.tolist()
        Unit_Price['In amount due to unit price (%)'] = ""
        pd.options.mode.chained_assignment = None

        for index in Unit_Price.index:
            GR_amt = Unit_Price[columns[18]][index]
            GR_qty = Unit_Price[columns[12]][index]
            Unit_price = (GR_amt / GR_qty)
            Unit_Price['In amount due to unit price (%)'][index] = Unit_price

        #  Rename Columns
        Unit_Price = Unit_Price.rename(
            columns={columns[6]: "GR Amt.in loc.cur."})
        Unit_Price = Unit_Price.rename(
            columns={columns[7]: "GR Qty"})
        Unit_Price = Unit_Price.rename(
            columns={columns[8]: "Unit Price"})
        Unit_Price = Unit_Price.rename(
            columns={columns[9]: "GR Amt.in loc.cur."})
        Unit_Price = Unit_Price.rename(
            columns={columns[10]: "GR Qty"})
        Unit_Price = Unit_Price.rename(
            columns={columns[11]: "Unit Price"})
        Unit_Price = Unit_Price.drop(columns=["index"])

        Unit_Price.to_excel(in_config["Unit_Path"], sheet_name=in_config["Unit_Sheet"], startrow=2, index=False)

        wb = load_workbook(in_config["Unit_Path"])
        ws = wb[in_config["Unit_Sheet"]]

        cell = ws['F2']
        cell.value = 'Current Quarter Q4'
        ws.merge_cells('F2:H2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['I2']
        cell.value = 'Previous Quarter Q3'
        ws.merge_cells('I2:K2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = ws['L2']
        cell.value = 'Increase/Decrease In Amount'
        ws.merge_cells('L2:S2')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        ws['F2'].font = font_style
        ws['I2'].font = font_style
        ws['L2'].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        ws['F2'].fill = fill_pattern
        ws['L2'].fill = fill_pattern
        fill_pattern = PatternFill(patternType="solid", fgColor="FFFF00")
        ws['I2'].fill = fill_pattern

        format_fill = PatternFill(patternType='solid', fgColor='ADD8E6')
        for c in ascii_lowercase:
            ws[c + "3"].fill = format_fill
            if c == 's':
                break
        m_row = ws.max_row

        ws.auto_filter.ref = "A3:S" + str(m_row)

        for cell in ws["E"]:
            cell.number_format = "#,###"
        for cell in ws["F"]:
            cell.number_format = "#,###"
        for cell in ws["G"]:
            cell.number_format = "#,###"
        for cell in ws["H"]:
            cell.number_format = "#,###"
        for cell in ws["I"]:
            cell.number_format = "#,###"
        for cell in ws["J"]:
            cell.number_format = "#,###"
        for cell in ws["K"]:
            cell.number_format = "#,###"
        for cell in ws["L"]:
            cell.number_format = "#,###"
        for cell in ws["M"]:
            cell.number_format = "#,###"
        for cell in ws["N"]:
            cell.number_format = "0"
        for cell in ws["O"]:
            cell.number_format = numbers.FORMAT_PERCENTAGE
        for cell in ws["P"]:
            cell.number_format = "0.00"
        for cell in ws["Q"]:
            cell.number_format = "0.00"
        for cell in ws["R"]:
            cell.number_format = "0.00"
        for cell in ws["S"]:
            cell.number_format = "0.00%"
        ws.delete_rows(m_row)
        ws.delete_rows(m_row - 1)

        ws['F1'] = '=SUBTOTAL(9,F4:F' + str(m_row) + ')'
        ws['G1'] = '=SUBTOTAL(9,G4:G' + str(m_row) + ')'
        ws['I1'] = '=SUBTOTAL(9,I4:I' + str(m_row) + ')'
        ws['J1'] = '=SUBTOTAL(9,J4:J' + str(m_row) + ')'
        ws['M1'] = '=SUBTOTAL(9,M4:M' + str(m_row) + ')'
        ws['L1'] = '=SUBTOTAL(9,L4:L' + str(m_row) + ')'
        ws['P1'] = '=SUBTOTAL(9,P4:P' + str(m_row) + ')'
        ws['R1'] = '=SUBTOTAL(9,R4:R' + str(m_row) + ')'

        # Auto-fit column width
        for c in ascii_lowercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 's':
                break

        wb.save(in_config["Unit_Path"])
        wb.close()

        return Unit_Price
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", end="")
        return notfound_error
    except BusinessException as business_error:
        print("Unit Price Comparison Process-", end="")
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", end="")
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", end="")
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", end="")
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Unit Price Comparison Process-", end="")
        return key_error


config = {}

if __name__ == "__main__":
    print(create_unit_price(config))


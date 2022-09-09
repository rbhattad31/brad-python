from string import ascii_uppercase

import pandas as pd
import numpy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from win32com import client
import pywintypes
from openpyxl.utils import get_column_letter


class BusinessException(Exception):
    pass


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


def same_mat_dvp(in_config):
    try:
        excel_file = pd.read_excel(in_config["Excel_file"], sheet_name=in_config["Q4sheet"], skiprows=6)

        input_file = excel_file[excel_file["Material No."].notna()]
        unit_price = excel_file[excel_file["unit price"].notna()]

        if excel_file.shape[0] == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Sheet is empty")
        elif len(unit_price) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Unit_price"],
                      body=in_config["Unit_price_body"])
            raise BusinessException("unit price column is empty")
        elif len(input_file) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Document is empty"],
                      body=in_config["Doc_body"])
            raise BusinessException("file is empty")
        else:
            pass
        max_pivot = pd.pivot_table(excel_file, index=["Material No.", "Material Desc", "Vendor Name"],
                                   values=["unit price"], aggfunc=numpy.max, margins=True, margins_name="Grand Total")

        max_pivot = max_pivot.reset_index()
        # print(max_pivot)
        min_pivot = pd.pivot_table(excel_file, index=["Material No.", "Material Desc", "Vendor Name"],
                                   values=["unit price"], aggfunc=numpy.min, margins=True, margins_name="Grand Total")
        min_pivot = min_pivot.reset_index()
        # print(min_pivot)

        com_file = pd.merge(max_pivot, min_pivot, how="outer",
                            on=["Material No.", "Material Desc", "Vendor Name"])
        com_file = com_file.replace(numpy.nan, '', regex=True)

        com_file["Deference"] = com_file["unit price_x"] - com_file["unit price_y"]
        com_file.drop(com_file[com_file["Deference"] <= 1].index, inplace=True)

        com_file = com_file.replace(numpy.nan, 0, regex=True)
        # print(com_file)

        com_file["Variance"] = com_file["Deference"] / com_file["unit price_y"]



        col_name = com_file.columns.values.tolist()

        com_file.sort_values(by=col_name[5], axis=0, ascending=False, inplace=True)
        # com_file[col_name[4]].values[-1] = grand_total
        com_file = com_file.rename(columns={col_name[3]: "Max of uint price"})
        com_file = com_file.rename(columns={col_name[4]: "Min of uint price"})

        com_file.to_excel(in_config["Out_file"], sheet_name=in_config["out_sheet"], index=False, startrow=1)


        wb = load_workbook(in_config["Out_file"])
        ws = wb[in_config["out_sheet"]]

        for cell in ws['D']:
            cell.number_format = "#,###.##"
        for cell in ws['E']:
            cell.number_format = "#,###.##"
        for cell in ws['F']:
            cell.number_format = "#,###.##"
        for cell in ws['G']:
            cell.number_format = "##%"
        ws.delete_rows(idx=3)

        m_row = ws.max_row
        # ws.auto_filter.ref = ws.dimensions

        FullRange = "A2:" + get_column_letter(ws.max_column) \
                    + str(ws.max_row)
        ws.auto_filter.ref = FullRange

        font_style = Font(name="Cambria", size=13, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "2"].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "2"].fill = fill_pattern
            if c == 'G':
                break

        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 35


        wb.save(in_config["Out_file"])
        wb.close()

        return com_file
    except FileNotFoundError as f_error:
        print("sent a mail file not found")
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["File_N"],
                  body=in_config["File_N_body"])
        return f_error
    except NameError as n_error:
        print("Name Error")
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Name_E"],
                  body=in_config["Name_E_body"])
        return n_error
    except KeyError as k_error:
        print("KeyError")
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Key"],
                  body=in_config["Key_body"])
        return k_error
    except ValueError as v_error:
        print("ValueError")
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Value_E"],
                  body=in_config["Value_E_body"])
        return v_error
    except SyntaxError as s_error:
        return s_error


config = {}
if __name__ == "__main__":
    print(same_mat_dvp(config))



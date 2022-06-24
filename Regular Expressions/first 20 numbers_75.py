from openpyxl import *
import re


def binary_to_decimal(binary):
    decimal = 0
    for digit in binary:
        decimal = decimal*2 + int(digit)
    return decimal


def oct_to_dec(octal):
    decimal_value = int(octal, 8)
    return decimal_value


def hex_to_dec(hex_val):
    decimal_value = int(hex_val, 16)
    return decimal_value


def check_binary(row_no, format_str, current_list_1):
    if re.match(r'[01]+', str(format_str)):
        current_val = binary_to_decimal(str(format_str))
        if current_list_1[0] == current_val:
            print('row {} is binary'.format(row_no))
        else:
            raise ValueError
    else:
        raise ValueError


def check_octal(row_no, format_str, current_list_1):
    if re.match(r'[0-7]+', str(format_str)):
        current_val = oct_to_dec(str(format_str))
        if current_list_1[0] == current_val:
            print('row {} is octal'.format(row_no))
        else:
            raise ValueError
    else:
        raise ValueError


def check_hex(row_no, format_str, current_list_1):
    if re.match(r'^[\da-fA-F]+', str(format_str)):
        current_val = hex_to_dec(str(format_str))
        if current_list_1[0] == current_val:
            print('row {} is hexadecimal'.format(row_no))
        else:
            raise ValueError
    else:
        raise ValueError


wb = load_workbook("sampleFile.xlsx")
ws = wb.active
m_r = ws.max_row
m_c = ws.max_column
data_list = []
for value in ws.iter_rows(min_row=1, max_row=m_r, min_col=1, max_col=m_c, values_only=True):
    data_list.append(value)
for i in range(1, len(data_list)):
    current_list = data_list[i]
    str_1 = current_list[1]

    try:
        check_binary(i, int(str_1), current_list)
    except ValueError:
        try:
            check_hex(i, str_1, current_list)
        except ValueError as e:
            try:
                check_octal(i, int(str_1), current_list)
            except ValueError:
                print('row {} is invalid'.format(i))
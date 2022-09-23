from string import ascii_uppercase
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import PatternFill, Font, Side, Border
from win32com import client
import pywintypes
from openpyxl.utils import get_column_letter


class BusinessException(Exception):
    pass


def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connections")
        return message_error
    except Exception as error:
        return error


def con_vendor_wise(in_config):
    try:
        read_present_quarter_pd = pd.read_excel(in_config["Input_file"], sheet_name=in_config["Sheet_name"])

        present_quarter_columns = read_present_quarter_pd.columns
        if in_config["purchase_register_1st_column_name"] in present_quarter_columns and \
                in_config["purchase_register_2nd_column_name"] in present_quarter_columns:
            print("The data is starting from first row only")
            pass

        else:
            print("The data is not starting from first row ")
            for index, row in read_present_quarter_pd.iterrows():

                if row[0] != in_config["purchase_register_1st_column_name"]:
                    read_present_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_present_quarter_pd.iloc[0]
            read_present_quarter_pd = read_present_quarter_pd[1:]
            read_present_quarter_pd.columns = new_header
            read_present_quarter_pd.reset_index(drop=True, inplace=True)
            read_present_quarter_pd.columns.name = None

            read_present_quarter_pd = read_present_quarter_pd[["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]]
            read_present_quarter_pd.to_excel(in_config["OutputFile_path"], sheet_name=in_config["Output_sheetName"],
                                             index=False)
            read_present_quarter_pd = pd.read_excel(in_config["OutputFile_path"],
                                                    sheet_name=in_config["Output_sheetName"])

        print(read_present_quarter_pd)
        if read_present_quarter_pd['GR Amt.in loc.cur.'].dtype == 'float64':
            print("Type is float")
        else:
            print("Column format not correct")
            raise TypeError

        Amount = read_present_quarter_pd[read_present_quarter_pd["GR Amt.in loc.cur."].notna()]
        Vendor_no = read_present_quarter_pd[read_present_quarter_pd["Vendor No."].notna()]

        if read_present_quarter_pd.shape[0] == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Sourcefile_subject"],
                      body=in_config["Body_mail1"])
            raise BusinessException("Sheet is empty")
        elif len(Amount) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Gr_amount"],
                      body=in_config["Gr_amount_body"])
            raise BusinessException("Empty column")

        elif len(Vendor_no) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Vendor_No"],
                      body=in_config["Vendor_No_body"])
            raise BusinessException("Vendor No column  missed")

        else:
            pass

        con_vendor_sheet = pd.pivot_table(read_present_quarter_pd, index=["Vendor No.", "Vendor Name"],
                                          values="GR Amt.in loc.cur.", aggfunc=numpy.sum,
                                          margins=True, margins_name="Grand Total", sort=True)
        con_vendor_sheet = con_vendor_sheet.reset_index()
        print(con_vendor_sheet)

        con_vendor_sheet = con_vendor_sheet.replace(numpy.nan, 0, regex=True)
        name_of_column = con_vendor_sheet.columns.values.tolist()
        con_vendor_sheet.drop(con_vendor_sheet.index[(con_vendor_sheet[name_of_column[2]] <= 0)], inplace=True)
        con_vendor_sheet = con_vendor_sheet.replace(numpy.nan, '', regex=True)

        grand_total = con_vendor_sheet[name_of_column[2]].values[-1]
        con_vendor_sheet[name_of_column[2]].values[-1] = 0
        con_vendor_sheet.sort_values(by=name_of_column[2], axis=0, ascending=False, inplace=True)

        con_vendor_sheet["Percentage"] = ""
        pd.options.mode.chained_assignment = None
        for index in con_vendor_sheet.index:
            vendor_amount = con_vendor_sheet[name_of_column[2]][index]

            if vendor_amount == 0:
                percentage = 1
            else:
                percentage = vendor_amount / grand_total
            con_vendor_sheet["Percentage"][index] = percentage
        con_vendor_sheet[name_of_column[2]].values[-1] = grand_total
        con_vendor_sheet = con_vendor_sheet.rename(columns={name_of_column[2]: in_config["Quarter_column_name"]})

        con_vendor_sheet.to_excel(in_config["OutputFile_path"], sheet_name=in_config["Output_sheetName"], index=False,
                                  startrow=17)
        print(con_vendor_sheet)

        wb = openpyxl.load_workbook(in_config["OutputFile_path"])
        ws = wb[in_config["Output_sheetName"]]

        for cell in ws["C"]:
            cell.number_format = "#,###.##"
        for cell in ws["D"]:
            cell.number_format = "##.##%"


        Full_range = "A18:" + get_column_letter(ws.max_column) \
                     + str(ws.max_row)

        ws.auto_filter.ref = Full_range

        font_style = Font(name="Cambria", size=13, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "18"].font = font_style
        for c in ascii_uppercase:
            ws[c + str(ws.max_row)].font = font_style
        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "18"].fill = fill_pattern
            if c == "D":
                break

        for c in ascii_uppercase:
            column_length = max(len(str(cell.value)) for cell in ws[c])
            ws.column_dimensions[c].width = column_length * 1.25
            if c == 'D':
                break

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=18, min_col=1, max_row=ws.max_row, max_col=4):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Cell merge for headers implementation
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws.merge_cells('A4:F4')
        ws.merge_cells('A5:F5')
        ws.merge_cells('A6:F6')
        ws.merge_cells('A7:F7')
        ws.merge_cells('A8:F8')
        ws.merge_cells('A9:F9')
        ws.merge_cells('A10:F10')
        ws.merge_cells('A11:F11')
        ws.merge_cells('A12:F12')
        ws.merge_cells('A13:F13')
        ws.merge_cells('A14:F14')

        # Headers implementation
        ws['A1'] = in_config['A1']
        ws['A2'] = in_config['A2']
        ws['A3'] = in_config['A3']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False

        wb.save(in_config["OutputFile_path"])
        wb.close()

    except SyntaxError as s_error:
        print("SyntaxError" + str(s_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Syn1"],
                  body=in_config["Synody_1"])
        return s_error
    except FileNotFoundError as f_error:
        print("FileNotFoundError" + str(f_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["File_N1"],
                  body=in_config["File_N1_body"])
        return f_error
    except NameError as n_error:
        print("NameError" + str(n_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Name_E1"],
                  body=in_config["Name_E1_body"])
        return n_error
    except KeyError as k_error:
        print("KeyError" + str(k_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Key1"],
                  body=in_config["Key1_body"])
        return k_error
    except ValueError as v_error:
        print("ValueError" + str(v_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Value_E1"],
                  body=in_config["Value_E1_body"])
        return v_error
    except AttributeError as a_error:
        print("AttributeError" + str(a_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["AttributeError1"],
                  body=in_config["AttributeError1_body"])
        return a_error
    except TypeError as t_error:
        print("TypeError" + str(t_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Type_E1"],
                  body=in_config["Type_E1_body"])
        return t_error
    except PermissionError as p_error:
        print("PermissionError" + str(p_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Perm_E1"],
                  body=in_config["Perm_E1_body"])
        return p_error
    except (ImportError, MemoryError, RuntimeError, Exception) as error:
        print("SystemError" + str(error))
        return error


config = {}

if __name__ == "__main__":
    print(con_vendor_wise(config))



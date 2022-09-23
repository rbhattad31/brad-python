import warnings
import pandas as pd
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from string import ascii_uppercase

from win32com import client
import pywintypes

warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)


class BusinessException(Exception):
    pass


# Send Outlook Mails
def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


# Defining a Function
def Create_Vendor_Wise(in_config):
    try:
        read_present_quarter_pd = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Present_quarter_sheet"])
        # print(len(read_present_quarter_pd))
        # print(read_present_quarter_pd.head())
        present_quarter_columns = read_present_quarter_pd.columns
        if in_config["purchase_register_1st_column_name"] in present_quarter_columns and in_config["purchase_register_2nd_column_name"] in present_quarter_columns:
            pass
        else:
            for index, row in read_present_quarter_pd.iterrows():
                if row[0] != in_config["purchase_register_1st_column_name"]:
                    read_present_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_present_quarter_pd.iloc[0]
            read_present_quarter_pd = read_present_quarter_pd[1:]
            read_present_quarter_pd.columns = new_header
            read_present_quarter_pd.reset_index(drop=True, inplace=True)
            read_present_quarter_pd.columns.name = None

        read_present_quarter_pd = read_present_quarter_pd[["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]]
        read_present_quarter_pd.to_excel(in_config["Output_Path"], sheet_name='q4', index=False)
        read_present_quarter_pd = pd.read_excel(in_config["Output_Path"], sheet_name='q4')
        # print(len(read_present_quarter_pd))
        # print(read_present_quarter_pd.head())

        read_previous_quarter_pd = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Previous_quarter_sheet"])
        # print(len(read_previous_quarter_pd))
        # vprint(read_previous_quarter_pd.head())

        previous_quarter_columns = read_previous_quarter_pd.columns
        if in_config["purchase_register_1st_column_name"] in previous_quarter_columns and in_config["purchase_register_2nd_column_name"] in previous_quarter_columns:
            pass
        else:
            for index, row in read_previous_quarter_pd.iterrows():
                if row[0] != in_config["purchase_register_1st_column_name"]:
                    read_previous_quarter_pd.drop(index, axis=0, inplace=True)
                else:
                    break
            new_header = read_previous_quarter_pd.iloc[0]
            read_previous_quarter_pd = read_previous_quarter_pd[1:]
            read_previous_quarter_pd.columns = new_header
            read_previous_quarter_pd.reset_index(level=0, drop=True, inplace=True)
            read_previous_quarter_pd.columns.name = None
        read_previous_quarter_pd = read_previous_quarter_pd[["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]]
        read_previous_quarter_pd.to_excel(in_config["Output_Path"], sheet_name='q3', index=False)
        read_previous_quarter_pd = pd.read_excel(in_config["Output_Path"], sheet_name='q3')
        # print(len(read_previous_quarter_pd))
        # print(read_previous_quarter_pd.head())

        # Checking Exception starts here
        # present quarter
        if read_present_quarter_pd.shape[0] == 0:
            send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Input Sheet Data is empty")

        present_quarter_columns_list = read_present_quarter_pd.columns.values.tolist()
        for col in ["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]:
            if col not in present_quarter_columns_list:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        Vendor_No_pd = read_present_quarter_pd[read_present_quarter_pd['Vendor No.'].notna()]
        Vendor_Name_pd = read_present_quarter_pd[read_present_quarter_pd['Vendor Name'].notna()]
        Gr_Amt_pd = read_present_quarter_pd[read_present_quarter_pd['GR Amt.in loc.cur.'].notna()]

        if len(Vendor_No_pd) == 0:
            send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"],
                      subject=in_config["Vendor No._subject"],
                      body=in_config["Vendor No._Body"])
            raise BusinessException("Vendor No. Column is empty")
        elif len(Vendor_Name_pd) == 0:
            send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"],
                      subject=in_config["Vendor Name_Subject"],
                      body=in_config["Vendor Name_Body"])
            raise BusinessException("Vendor Name Column is empty")
        elif len(Gr_Amt_pd) == 0:
            send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # present quarter exceptions ends here
        # previous quarter exceptions starts here
        if read_previous_quarter_pd.shape[0] == 0:
            send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Input Sheet Data is empty")

        previous_quarter_columns_list = read_previous_quarter_pd.columns.values.tolist()
        for col in ["Vendor No.", "Vendor Name", "GR Amt.in loc.cur."]:
            if col not in previous_quarter_columns_list:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows

        Vendor_No_pd = read_previous_quarter_pd[read_previous_quarter_pd['Vendor No.'].notna()]
        Vendor_Name_pd = read_previous_quarter_pd[read_previous_quarter_pd['Vendor Name'].notna()]
        Gr_Amt_pd = read_previous_quarter_pd[read_previous_quarter_pd['GR Amt.in loc.cur.'].notna()]


        if len(Vendor_No_pd) == 0:

            send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"],
                      subject=in_config["Vendor No._subject"],
                      body=in_config["Vendor No._Body"])
            raise BusinessException("Vendor No. Column is empty")


        elif len(Vendor_Name_pd) == 0:
            send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"],
                      subject=in_config["Vendor Name_Subject"],
                      body=in_config["Vendor Name_Body"])
            raise BusinessException("Vendor Name Column is empty")

        elif len(Gr_Amt_pd) == 0:
            send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # exception ends here


        # create pivot table

        Vendor_wise_pd = pd.pivot_table(read_present_quarter_pd, index=["Vendor No.", "Vendor Name"],
                                        values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                        margins_name="Grand Total")
        Vendor_wise_pd = Vendor_wise_pd[:-1]
        # reset "indices created during pivot table creation" - for merging
        Vendor_wise_pd = Vendor_wise_pd.reset_index()
        # read previous quarters final working file - pd will be replaced with Nan in any blank cells
        previous_quarter_final_file_pd = pd.pivot_table(read_previous_quarter_pd,
                                                        index=['Vendor No.', 'Vendor Name'],
                                                        values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                                        margins_name="Grand Total")
        previous_quarter_final_file_pd = previous_quarter_final_file_pd[:-1]
        # replace Nan with blank
        previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, 0, regex=True)

        # merging present and previous quarter purchase type wise data -  pd will be replaced with Nan in any blank cells
        merge_pd = pd.merge(Vendor_wise_pd, previous_quarter_final_file_pd, how="outer",
                            on=["Vendor No.", "Vendor Name"])

        # replacing all Nan's with zeros in Present and previous Quarter's values columns
        merge_pd = merge_pd.replace(numpy.nan, 0, regex=True)

        Col_List = merge_pd.columns.values.tolist()
        # returns as ['Valuation Class', 'Valuation Class Text', 'GR Amt.in loc.cur.', 'Previous Quarter']

        # dropping columns present and previous quarters both have values as zero
        merge_pd.drop(merge_pd.index[(merge_pd[Col_List[2]] == 0) & (merge_pd[Col_List[3]] == 0)],
                      inplace=True)
        merge_pd.sort_values(by=Col_List[2], axis=0, ascending=False, inplace=True)

        # create a new column - Success
        merge_pd['Variance'] = 0

        pd.options.mode.chained_assignment = None

        # variance formula implementation using index
        for index in merge_pd.index:
            Present_quarter_row_value = merge_pd[Col_List[2]][index]
            Previous_quarter_row_value = merge_pd[Col_List[3]][index]
            variance = Present_quarter_row_value - Previous_quarter_row_value
            merge_pd['Variance'][index] = variance

        Col_List = merge_pd.columns.values.tolist()
        merge_pd.drop(merge_pd.index[(merge_pd[Col_List[3]] == 0) & (merge_pd[Col_List[4]] == 0)],
                      inplace=True)
        merge_pd['Percentage'] = ''
        pd.options.mode.chained_assignment = None
        # variance formula implementation using index
        for index in merge_pd.index:
            Previous_quarter_row_value = merge_pd[Col_List[3]][index]
            Variance_row_value = merge_pd[Col_List[4]][index]
            if Previous_quarter_row_value == 0:
                Percentage = 1
            elif Variance_row_value == 0:
                Percentage = 1
            else:
                Percentage = Variance_row_value / Previous_quarter_row_value
            merge_pd['Percentage'][index] = Percentage
        Vendor_wise_comparatives_pd = merge_pd.rename(
            columns={Col_List[2]: in_config["Present_quarter_Column_Name"]})
        Vendor_wise_comparatives_pd = Vendor_wise_comparatives_pd.rename(
            columns={Col_List[3]: in_config["Previous_quarter_Column_Name"]})

        present_quarter_subtotal = Vendor_wise_comparatives_pd[in_config["Present_quarter_Column_Name"]].sum()
        print(present_quarter_subtotal)
        previous_quarter_subtotal = Vendor_wise_comparatives_pd[in_config["Previous_quarter_Column_Name"]].sum()
        print(previous_quarter_subtotal)
        variance_subtotal = present_quarter_subtotal - previous_quarter_subtotal
        print(variance_subtotal)

        Vendor_wise_comparatives_pd.to_excel(in_config["Output_Path"],
                                             sheet_name=in_config["Vendor_Wise_Sheet_Name"], startrow=17,
                                             index=False)

        wb = openpyxl.load_workbook(in_config["Output_Path"])
        ws = wb[in_config["Vendor_Wise_Sheet_Name"]]

        m_row = ws.max_row

        ws['C17'] = '=SUBTOTAL(9,C19:C'+str(m_row)+')'
        ws['D17'] = '=SUBTOTAL(9,D19:D'+str(m_row)+')'
        ws['E17'] = '=SUBTOTAL(9,E19:E'+str(m_row)+')'

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "17"].font = font_style

        m_row = ws.max_row
        for c in ascii_uppercase:
            ws[c + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "18"].fill = fill_pattern
            if c == 'F':
                break
        m_row = ws.max_row
        # print("A2:E" + str(m_row))
        ws.auto_filter.ref = "A18:F" + str(m_row)
        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 20
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["B"].width = 35
        ws.delete_rows(m_row)
        ws.delete_rows(m_row - 1)

        ws['C1'] = present_quarter_subtotal
        ws['D1'] = previous_quarter_subtotal
        ws['E1'] = variance_subtotal
        for cell in ws["C"]:
            cell.number_format = "#,##,###"
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='center')

        for cell in ws["D"]:
            cell.number_format = "#,##,###"
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='center')
        for cell in ws['E']:
            cell.number_format = '##,##'
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='center')

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=18, min_col=1, max_row=ws.max_row, max_col=6):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Cell merge for headers implementation
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws.merge_cells('A4:F4')
        ws.merge_cells('A5:F5')
        ws.merge_cells('A6:F6')
        ws.merge_cells('A7:F7')
        ws.merge_cells('A8:F8')
        ws.merge_cells('A9:F9')
        ws.merge_cells('A10:F10')
        ws.merge_cells('A11:F11')
        ws.merge_cells('A12:F12')
        ws.merge_cells('A13:F13')
        ws.merge_cells('A14:F14')

        # Headers implementation
        ws['A1'] = in_config['A1']
        ws['A2'] = in_config['A2']
        ws['A3'] = in_config['A3']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False

        wb.save(in_config["Output_Path"])

        return Vendor_wise_comparatives_pd

    # Excepting Errors here
    except PermissionError as file_error:
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=in_config["SystemE_Subject"],
                  body=in_config["SystemE_Body"])
        print("Please close the file")
        return file_error
    except FileNotFoundError as notfound_error:
        subject = in_config["FileNotFound_Subject"]
        body = in_config["FileNotFound_Body"]
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Vendor Wise Comparitives Process", end="")
        return notfound_error
    except BusinessException as business_error:
        print("Vendor Wise Comparitives Process-", end="")
        return business_error
    except ValueError as value_error:
        subject = in_config["SheetMiss_Subject"]
        body = in_config["SheetMiss_Body"]
        body = body.replace("ValueError +", str(value_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Vendor Wise Comparitives Process-", end="")
        return value_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Vendor Wise Comparitives Process-", end="")
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Vendor Wise Comparitives Process-", end="")
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["To_Address"], cc=in_config["CC_Address"], subject=subject, body=body)
        print("Vendor Wise Comparitives Process-", end="")
        return key_error


config = {}

if __name__ == "__main__":
    print(Create_Vendor_Wise(config))


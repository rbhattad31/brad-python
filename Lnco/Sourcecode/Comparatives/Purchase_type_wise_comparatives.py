# Importing Libraries

import pandas as pd
import numpy

import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from string import ascii_uppercase

from win32com import client
import pywintypes


class BusinessException(Exception):
    pass


# Send Outlook Mails
def send_mail(to, cc, subject, body):
    try:
        outlook = client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.cc = cc
        mail.Subject = subject
        mail.Body = body
        mail.Send()
    except pywintypes.com_error as message_error:
        print("Sendmail error - Please check outlook connection")
        return message_error
    except Exception as error:
        return error


# Defining a Function
def create_purchase_type_wise(in_config):
    try:
        read_excel_pd = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q4 Sheet"], skiprows=6)
        read_excel_pd_2 = pd.read_excel(in_config["ExcelPath"], sheet_name=in_config["Q3 Sheet_Purchase"], skiprows=5)
        # Check Exception
        if read_excel_pd.shape[0] == 0 or read_excel_pd_2.shape[0] == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["subject_mail"],
                      body=in_config["Body_mail"])
            raise BusinessException("Input Sheet Data is empty")

        Q3Sheet_col = read_excel_pd_2.columns.values.tolist()
        for col in ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur."]:
            if col not in Q3Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        Q4Sheet_col = read_excel_pd.columns.values.tolist()
        for col in ["Valuation Class", "Valuation Class Text", "GR Amt.in loc.cur."]:
            if col not in Q4Sheet_col:
                subject = in_config["ColumnMiss_Subject"]
                body = in_config["ColumnMiss_Body"]
                body = body.replace("ColumnName +", col)

                send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
                raise BusinessException(col + " Column is missing")

        # Filter Rows
        Valuation_pd = read_excel_pd[read_excel_pd['Valuation Class'].notna()]
        Valuation_Text_pd = read_excel_pd[read_excel_pd['Valuation Class Text'].notna()]
        Gr_Amt_pd = read_excel_pd[read_excel_pd['GR Amt.in loc.cur.'].notna()]

        if len(Valuation_pd) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Valuation Class_subject"],
                      body=in_config["Valuation Class_Body"])
            raise BusinessException("Valuation Class Column is empty")

        elif len(Valuation_Text_pd) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Valuation Class Text_Subject"],
                      body=in_config["Valuation Class Text_Body"])
            raise BusinessException("Valuation Class Text Column is empty")

        elif len(Gr_Amt_pd) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        Valuation_pd_2 = read_excel_pd_2[read_excel_pd_2['Valuation Class'].notna()]
        Valuation_Text_pd_2 = read_excel_pd_2[read_excel_pd_2['Valuation Class Text'].notna()]
        Gr_Amt_pd_2 = read_excel_pd_2[read_excel_pd_2['GR Amt.in loc.cur.'].notna()]

        if len(Valuation_pd_2) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Valuation Class_subject"],
                      body=in_config["Valuation Class_Body"])
            raise BusinessException("Valuation Class Column is empty")

        elif len(Valuation_Text_pd_2) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Valuation Class Text_Subject"],
                      body=in_config["Valuation Class Text_Body"])
            raise BusinessException("Valuation Class Text Column is empty")

        elif len(Gr_Amt_pd_2) == 0:
            send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["Gr Amt_Subject"],
                      body=in_config["Gr Amt_Body"])
            raise BusinessException("GR Amt Column is empty")
        else:
            pass

        # create pivot table

        purchase_type_wise_pd = pd.pivot_table(read_excel_pd, index=["Valuation Class", "Valuation Class Text"],
                                               values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                               margins_name="Grand Total")

        # reset "indices created during pivot table creation" - for merging
        purchase_type_wise_pd = purchase_type_wise_pd.reset_index()



        # read previous quarters final working file - pd will be replaced with Nan in any blank cells
        previous_quarter_final_file_pd = pd.pivot_table(read_excel_pd_2, index=["Valuation Class", "Valuation Class Text"],
                                               values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                               margins_name="Grand Total")

        # replace Nan with blank
        previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, '', regex=True)

        # merging present and previous quarter purchase type wise data -  pd will be replaced with Nan in any blank cells
        merge_pd = pd.merge(purchase_type_wise_pd, previous_quarter_final_file_pd, how="outer",
                            on=["Valuation Class", "Valuation Class Text"])

        # replacing all Nan's with zeros in Present and previous Quarter's values columns
        merge_pd = merge_pd.replace(numpy.nan, 0, regex=True)

        Col_List = merge_pd.columns.values.tolist()
        # returns as ['Valuation Class', 'Valuation Class Text', 'GR Amt.in loc.cur.', 'Previous Quarter']

        # dropping columns present and previous quarters both have values as zero
        merge_pd.drop(merge_pd.index[(merge_pd[Col_List[2]] == 0) & (merge_pd[Col_List[3]] == 0)],
                      inplace=True)

        # create a new column - Success
        merge_pd['Variance'] = 0

        pd.options.mode.chained_assignment = None

        # variance formula implementation using index
        for index in merge_pd.index:
            Q4 = merge_pd[Col_List[2]][index]
            Q3 = merge_pd[Col_List[3]][index]
            if Q3 == 0:
                variance = 1
            else:
                variance = (Q4 - Q3) / Q3
            merge_pd['Variance'][index] = variance

        # copy present quarter Amount column Grand total, set it as zero, sort the data frame and reassign the value.
        grand_total = merge_pd[Col_List[2]].values[-1]
        merge_pd[Col_List[2]].values[-1] = 0

        merge_pd[Col_List[2]].values[-1] = grand_total

        purchase_type_wise_comparatives_pd = merge_pd.rename(
            columns={Col_List[2]: in_config["Q4 Column"]})
        purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.rename(
            columns={Col_List[3]: in_config["Q3 Column"]})

        purchase_type_wise_comparatives_pd.to_excel(in_config["Type_Path"],
                                                    sheet_name=in_config["Purchase Type Wise"],
                                                    index=False, startrow=16)

        wb = openpyxl.load_workbook(in_config["Type_Path"])
        ws = wb[in_config["Purchase Type Wise"]]

        for cell in ws['C']:
            cell.number_format = '#,###.##'
        for cell in ws['D']:
            cell.number_format = '#,###.##'
        for cell in ws['E']:
            cell.number_format = '0.0%'

        font_style = Font(name="Cambria", size=12, bold=True, color="000000")
        for c in ascii_uppercase:
            ws[c + "17"].font = font_style

        m_row = ws.max_row
        for c in ascii_uppercase:
            ws[c + str(m_row)].font = font_style

        fill_pattern = PatternFill(patternType="solid", fgColor="ADD8E6")
        for c in ascii_uppercase:
            ws[c + "17"].fill = fill_pattern
            if c == 'E':
                break

        for c in ascii_uppercase:
            ws.column_dimensions[c].width = 25
        ws.column_dimensions["E"].width = 15

        thin = Side(border_style="thin", color='b1c5e7')

        for row in ws.iter_rows(min_row=17, min_col=1, max_row=ws.max_row, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        font_style1 = Font(name='Cambria', size=12, color='002060', bold=False)
        font_style2 = Font(name='Cambria', size=12, color='002060', bold=True, underline='single')
        font_style3 = Font(name='Cambria', size=14, color='002060', bold=True)

        # Cell merge for headers implementation
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        ws.merge_cells('A4:E4')
        ws.merge_cells('A5:E5')
        ws.merge_cells('A6:E6')
        ws.merge_cells('A7:E7')
        ws.merge_cells('A8:E8')
        ws.merge_cells('A9:E9')
        ws.merge_cells('A10:E10')
        ws.merge_cells('A11:E11')
        ws.merge_cells('A12:E12')
        ws.merge_cells('A13:E13')
        ws.merge_cells('A14:E14')

        # Headers implementation
        ws['A1'] = in_config['A1']
        ws['A2'] = in_config['A2']
        ws['A3'] = in_config['A3']
        ws['A4'] = in_config['A4']
        ws['A5'] = in_config['A5']
        ws['A7'] = in_config['A7']
        ws['A8'] = in_config['A8']
        ws['A10'] = in_config['A10']
        ws['A11'] = in_config['A11']
        ws['A12'] = in_config['A12']

        # Headers formatting and styling
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=1):
            for cell in row:
                cell.font = font_style3

        for row in ws.iter_rows(min_row=7, min_col=1, max_row=7, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=10, min_col=1, max_row=10, max_col=1):
            for cell in row:
                cell.font = font_style2

        for row in ws.iter_rows(min_row=8, min_col=1, max_row=8, max_col=1):
            for cell in row:
                cell.font = font_style1

        for row in ws.iter_rows(min_row=11, min_col=1, max_row=12, max_col=1):
            for cell in row:
                cell.font = font_style1

        ws.sheet_view.showGridLines = False

        wb.save(in_config["Type_Path"])

        return purchase_type_wise_comparatives_pd

    # Excepting Errors here
    except FileNotFoundError as notfound_error:
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=in_config["subject_file_not_found"],
                  body=in_config["body_file_not_found"])
        print("Purchase Type Wise Comparatives Process-", end="")
        return notfound_error
    except ValueError as V_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(V_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", end="")
        return V_error
    except BusinessException as business_error:
        print("Purchase Type Wise Comparatives Process-", end="")
        return business_error
    except TypeError as type_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(type_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", end="")
        return type_error
    except (OSError, ImportError, MemoryError, RuntimeError, Exception) as error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", end="")
        return error
    except KeyError as key_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(key_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Purchase Type Wise Comparatives Process-", end="")
        return key_error
    except PermissionError as file_error:
        subject = in_config["SystemError_Subject"]
        body = in_config["SystemError_Body"]
        body = body.replace("SystemError +", str(file_error))
        send_mail(to=in_config["to_mail"], cc=in_config["cc_mail"], subject=subject, body=body)
        print("Please close the file")
        return file_error


config = {}

if __name__ == "__main__":
    print(create_purchase_type_wise(config))

